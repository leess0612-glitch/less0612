"""
LG 수수료 엑셀 → JSON 변환 파서 (독립 실행 버전)
에이컴즈: ★LG 26년 04월 구독전문점 수수료안내_0401.xlsx (시트: 정수기 수수료)
티엘:     2026.04.09 수수료.xlsx (시트: LG정수기)

사용법: python parse_lg_excel.py
"""
import openpyxl
import json
import re
import os
from datetime import datetime


# ─────────────────────────────────────────────
# 경로 / 상수
# ─────────────────────────────────────────────

AC_PATH   = r'C:\Users\a\Documents\렌탈정책\26.04\★LG 26년 04월 구독전문점 수수료안내_0401.xlsx'
TL_PATH   = r'C:\Users\a\Documents\렌탈정책\26.04\2026.04.09 수수료.xlsx'
AC_SHEET  = '정수기 수수료'
TL_SHEET  = 'LG정수기'
DATA_ROW  = 4   # 1-indexed 데이터 시작 행 (양쪽 동일)

# 수수료 열 번호 (1-indexed)
AC_FEE_COL = 21  # U열
TL_FEE_COL = 14  # N열

# 렌탈주관사(LG전자) 별도 수수료 — 매달 확인 후 수정
# 에이컴즈·티엘 각각 다를 수 있음
EXTRA_FEE_AK = 100000
EXTRA_FEE_TL = 100000

# 팝업 예외: (정규화 모델코드, 결합여부) → 팝업 메시지
# WS511SH / WS513SH 결합 → 티엘 미운영
POPUP_EXCEPTIONS = {
    ('WS511SH', '결합'): '티엘 결합 접수 불가 — 에이컴즈로만 접수 가능\n확인 필요',
    ('WS513SH', '결합'): '티엘 결합 접수 불가 — 에이컴즈로만 접수 가능\n확인 필요',
}


# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────

def clean(v):
    if v is None:
        return ''
    return str(v).strip().replace('\xa0', ' ').replace('\u3000', ' ')


def to_int(v):
    try:
        return int(v) if v else 0
    except (ValueError, TypeError):
        return 0


def normalize_model_code(raw):
    """
    LG E열 모델코드 정규화.

    핵심 원칙:
      엑셀에서 같은 행에 묶인 경우(괄호 / 슬래시 / 콤마)만 동일 제품으로 간주하여 base 코드 추출.
      별도 행으로 분리된 standalone 모델은 수수료가 다를 수 있으므로 코드 그대로 유지.

    규칙:
      1. 개행(\\n) 이후 색상명 한글 제거
      2. 콤마(,) 이후 병기 모델 제거  ← 수수료 동일 변형이 콤마로 병기됨
      3. 괄호 (C/W/N/B) 또는 슬래시 WD525AHB/ACB/AGB/AS 형태
         → 명시적 색상 변형이 있는 경우에만 base 코드 추출
         → W[단일자][3자리숫자](SH|[단일자]) 패턴 적용
      4. standalone (괄호/슬래시 없음) → 코드 그대로 반환
         → WU523AS, WD523VC, WS511SH 등 별도 제품으로 유지

    예시:
      WD722R(K/H/E)              → WD722R   (괄호 제거 후 base 추출)
      WU923A(C/W/N/B)B, AS       → WU923A   (콤마+괄호 → base 추출)
      WU523A(C/W)B               → WU523A   (괄호 → base 추출)
      WD525AHB/ACB/AGB/AS        → WD525A   (슬래시 변형 → 첫 파트 base 추출)
      WU523AS  (별도 행)          → WU523AS  (standalone → 유지)
      WU823AS  (별도 행)          → WU823AS  (standalone → 유지)
      WD523VC  (별도 행)          → WD523VC  (standalone → 유지)
      WS511SH  (별도 행)          → WS511SH  (standalone → 유지)
    """
    if not raw:
        return ''
    # 1. 개행 이후 제거
    first_line = raw.split('\n')[0].strip()
    # 2. 콤마 이후 제거
    first_part = first_line.split(',')[0].strip()

    has_parens = bool(re.search(r'\([A-Z/]+\)', first_part))
    has_slash  = '/' in first_part and not has_parens

    if has_parens:
        # 3a. 괄호 안 색상코드 제거 후 base 추출
        cleaned = re.sub(r'\([A-Z/]+\)', '', first_part).replace(' ', '')
        m = re.match(r'^(W[A-Z]\d{3}(?:SH|[A-Z]))', cleaned)
        return m.group(1) if m else cleaned

    elif has_slash:
        # 3b. 슬래시 색상 변형: 첫 번째 파트에서 base 추출
        first_slash = first_part.split('/')[0].strip()
        m = re.match(r'^(W[A-Z]\d{3}(?:SH|[A-Z]))', first_slash)
        return m.group(1) if m else first_slash

    else:
        # 4. standalone → 코드 그대로 (색상 suffix 제거 안 함)
        return first_part.replace(' ', '')


def normalize_manage(j_val):
    """
    J열 관리주기 → (manage_type, manage_cycle)
      '자가'         → ('자가관리', None)
      '6개월'/'4개월'/'3개월' → ('방문관리', cycle_str)
    """
    s = clean(j_val)
    if s == '자가':
        return '자가관리', None
    if '개월' in s:
        return '방문관리', s
    return s or '알수없음', None


def normalize_promo(f_val, g_val, h_val):
    """F/G/H열 프로모션 정규화. G열 'X' → None."""
    f = clean(f_val).replace('\n', ' ').strip() or None
    g = clean(g_val)
    g = None if (not g or g == 'X') else g
    h = clean(h_val).replace('\n', ' ').strip() or None
    return {'monthlyDiscount': f, 'tasa': g, 'modelDiscount': h}


# ─────────────────────────────────────────────
# 시트 로더 (병합 셀 처리)
# ─────────────────────────────────────────────

def load_sheet_with_merges(wb, sheet_name):
    ws = wb[sheet_name]
    merge_map = {}
    for mr in ws.merged_cells.ranges:
        top_val = ws.cell(mr.min_row, mr.min_col).value
        for row in range(mr.min_row, mr.max_row + 1):
            for col in range(mr.min_col, mr.max_col + 1):
                merge_map[(row, col)] = top_val

    def get(r, c):
        return merge_map.get((r, c), ws.cell(r, c).value)

    rows = []
    for r in range(DATA_ROW, ws.max_row + 1):
        row = {c: get(r, c) for c in range(1, ws.max_column + 1)}
        rows.append(row)
    return rows


# ─────────────────────────────────────────────
# 행 파싱
# ─────────────────────────────────────────────

def parse_sheet_rows(rows, fee_col):
    """
    rows: load_sheet_with_merges 결과
    fee_col: 수수료 열 번호 (에이컴즈=21, 티엘=14)

    반환: {(약정년수, 모델코드, 관리주기_or_자가, 결합여부): row_info}
    """
    result = {}
    for row in rows:
        year        = to_int(row.get(1))   # A열: 약정년수
        lineup      = clean(row.get(3))    # C열: 라인업
        spec        = clean(row.get(4))    # D열: 스펙구분
        model_raw   = clean(row.get(5))    # E열: 모델명+색상
        promo_f     = row.get(6)           # F열: 월요금할인
        promo_g     = row.get(7)           # G열: 타사보상
        promo_h     = row.get(8)           # H열: 전용모델할인
        j_val       = row.get(10)          # J열: 관리주기
        combine     = clean(row.get(11))   # K열: 단품/결합
        comb_disc   = to_int(row.get(12))  # L열: 결합할인가
        monthly_fee = to_int(row.get(13))  # M열: 렌탈료
        commission  = to_int(row.get(fee_col))  # 수수료

        if not model_raw or year == 0:
            continue
        if monthly_fee == 0 and commission == 0:
            continue

        model_code = normalize_model_code(model_raw)
        manage_type, manage_cycle = normalize_manage(j_val)
        promo = normalize_promo(promo_f, promo_g, promo_h)

        # 키: (약정년수, 모델코드, 관리주기(없으면'자가'), 결합여부)
        key = (year, model_code, manage_cycle or '자가', combine)
        result[key] = {
            'year':          year,
            'modelCode':     model_code,
            'lineup':        lineup,
            'spec':          spec,
            'manageType':    manage_type,
            'manageCycle':   manage_cycle,
            'combineType':   combine,
            'combineDiscount': comb_disc,
            'monthlyFee':    monthly_fee,
            'commission':    commission,
            'promo':         promo,
        }
    return result


# ─────────────────────────────────────────────
# 제품 목록 생성 (에이컴즈 + 티엘 비교)
# ─────────────────────────────────────────────

def build_products(ac_data, tl_data):
    all_keys = set(ac_data.keys()) | set(tl_data.keys())

    # 모델코드별 키 그룹화
    model_groups = {}
    for key in all_keys:
        mc = key[1]
        model_groups.setdefault(mc, []).append(key)

    products = []
    for model_code in sorted(model_groups):
        keys = sorted(model_groups[model_code])
        sample = ac_data.get(keys[0]) or tl_data.get(keys[0], {})

        product = {
            'id':        model_code,
            'modelCode': model_code,
            'lineup':    sample.get('lineup', ''),
            'spec':      sample.get('spec', ''),
            'options':   [],
        }

        for key in keys:
            year, _, cycle_key, combine = key
            ac_row = ac_data.get(key)
            tl_row = tl_data.get(key)
            row    = ac_row or tl_row

            ak_fee     = (ac_row['commission'] + EXTRA_FEE_AK) if ac_row else 0
            tl_fee     = (tl_row['commission'] + EXTRA_FEE_TL) if tl_row else 0
            ak_monthly = ac_row['monthlyFee'] if ac_row else 0
            tl_monthly = tl_row['monthlyFee'] if tl_row else 0

            # 접수처 추천
            if ak_fee > 0 and tl_fee > 0:
                if ak_fee > tl_fee:
                    recommended = '에이컴즈'
                elif tl_fee > ak_fee:
                    recommended = '티엘'
                else:
                    recommended = '동일'
            elif ak_fee > 0:
                recommended = '에이컴즈'
            elif tl_fee > 0:
                recommended = '티엘'
            else:
                recommended = None

            # 팝업 예외 확인
            popup = POPUP_EXCEPTIONS.get((model_code, combine))

            product['options'].append({
                'contractYears':  year,
                'manageType':     row['manageType'],
                'manageCycle':    row.get('manageCycle'),
                'combineType':    combine,
                'combineDiscount': row.get('combineDiscount', 0),
                'monthlyFee':     ak_monthly or tl_monthly,
                'commission': {
                    'ak': ak_fee or None,
                    'tl': tl_fee or None,
                },
                'recommendedOffice': recommended,
                'popup':          popup,
                'promo':          row.get('promo', {}),
            })

        products.append(product)
    return products


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────

def parse_lg(ac_filepath=AC_PATH, tl_filepath=TL_PATH):
    print(f'[에이컴즈] 파싱 중: {os.path.basename(ac_filepath)}')
    ac_wb   = openpyxl.load_workbook(ac_filepath, data_only=True)
    ac_rows = load_sheet_with_merges(ac_wb, AC_SHEET)
    ac_data = parse_sheet_rows(ac_rows, AC_FEE_COL)

    print(f'[티엘]     파싱 중: {os.path.basename(tl_filepath)}')
    tl_wb   = openpyxl.load_workbook(tl_filepath, data_only=True)
    tl_rows = load_sheet_with_merges(tl_wb, TL_SHEET)
    tl_data = parse_sheet_rows(tl_rows, TL_FEE_COL)

    products = build_products(ac_data, tl_data)

    # 통계
    popup_count  = sum(1 for p in products for o in p['options'] if o.get('popup'))
    ak_only      = sum(1 for p in products for o in p['options']
                       if o['commission']['ak'] and not o['commission']['tl'])
    tl_only      = sum(1 for p in products for o in p['options']
                       if o['commission']['tl'] and not o['commission']['ak'])
    diff_count   = sum(1 for p in products for o in p['options']
                       if o['recommendedOffice'] not in ('동일', None))

    print(f'\n에이컴즈 항목: {len(ac_data)}개')
    print(f'티엘     항목: {len(tl_data)}개')
    print(f'제품 수:       {len(products)}개')
    print(f'팝업 예외:     {popup_count}건  ← WS511SH/WS513SH 결합')
    print(f'에이컴즈 전용: {ak_only}건')
    print(f'티엘 전용:     {tl_only}건')
    print(f'수수료 차이:   {diff_count}건')

    return {
        'metadata': {
            'brand':    'LG전자',
            'category': '정수기',
            'sourceFiles': {
                'ak': os.path.basename(ac_filepath),
                'tl': os.path.basename(tl_filepath),
            },
            'parsedAt': datetime.now().strftime('%Y-%m-%d %H:%M'),
        },
        'products': products,
    }


def inject_lg_into_html(data, base_dir):
    """
    렌탈수수료_*.html (가장 최신) 에서 __LG_DATA__ 플레이스홀더를 LG JSON으로 대체한다.
    """
    import glob as glob_mod

    # 가장 최신 출력 HTML 탐색 (렌탈수수료_2604.html 등)
    pattern = os.path.join(base_dir, '렌탈수수료_[0-9]*.html')
    candidates = sorted(glob_mod.glob(pattern))
    if not candidates:
        print('⚠️  렌탈수수료_*.html 파일을 찾지 못했습니다. parse_excel.py 를 먼저 실행하세요.')
        return

    target = candidates[-1]  # 가장 최신 파일
    with open(target, 'r', encoding='utf-8') as f:
        html = f.read()

    if '__LG_DATA__' not in html:
        print(f'⚠️  {os.path.basename(target)} 에 __LG_DATA__ 플레이스홀더가 없습니다.')
        return

    lg_js = json.dumps(data, ensure_ascii=False)
    html_out = html.replace('__LG_DATA__', lg_js)

    with open(target, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f'HTML 주입 완료: {target}')


if __name__ == '__main__':
    import sys
    ac = sys.argv[1] if len(sys.argv) > 1 else AC_PATH
    tl = sys.argv[2] if len(sys.argv) > 2 else TL_PATH

    data = parse_lg(ac, tl)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(base_dir, 'lg_data.json')
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'\nJSON 저장: {out}')
    print(f'완료! 제품 수: {len(data["products"])}개')

    inject_lg_into_html(data, base_dir)
