"""
쿠쿠 수수료 엑셀 → JSON 변환 파서
티엘: 2026.04.21 수수료.xlsx (시트: 쿠쿠)

열 구조:
  A: 상품군 (카테고리)
  B: 모델명 (색상별 코드 or 제품명 — 모델코드 도출 보조용)
  C: 상품명 (무시)
  D: 제품명(모델코드) — 주 키. 비어있으면 B열에서 추출
  E: 의무약정 (36M / 48M / 60M / 72M / 84M / 39M / 12M)
  F: 소유권 (무시)
  G: 프로모션 — 타사보상 or 반값 포함 행만 채택.
     단, 해당 모델에 타사보상/반값 행이 아예 없으면 빈칸 행도 포함.
  H: 구분 (일반 / 패키지 / 패키지10%)
  I: 방문주기 (4개월/6개월… → 방문관리, 12개월 → 셀프관리, 없음 → 관리없음)
  J: 렌탈료(월요금)
  K: 수수료 a
  L: 수수료 b  →  max(K, L) 사용 (한쪽이 '-'/None이면 나머지만)
"""
import openpyxl
import json
import re
import os
from datetime import datetime


# ─────────────────────────────────────────────
# 경로
# ─────────────────────────────────────────────
TL_PATH = r'C:\Users\a\Documents\렌탈정책\26.04\2026.04.21 수수료.xlsx'
SHEET   = '쿠쿠'
DATA_ROW = 4  # 4행부터 데이터


# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────
def clean(v):
    if v is None:
        return ''
    return str(v).strip().replace('\xa0', ' ').replace('\u3000', ' ')


def to_float(v):
    """'-' 또는 None → None, 그 외 float"""
    if v is None:
        return None
    s = str(v).strip()
    if s in ('-', ''):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_months(e_val):
    """E열 약정 → 월수 (36M → 36)"""
    m = re.match(r'^(\d+)M$', clean(e_val))
    return int(m.group(1)) if m else 0


def months_to_label(months):
    """월수 → 표기 레이블 (36 → '3년', 39 → '39개월')"""
    if months % 12 == 0:
        return f'{months // 12}년'
    return f'{months}개월'


def normalize_model_code(b_first_line, d_raw):
    """
    D열 우선. 비어있으면 B열 첫 줄에서 모델코드 추출.
    B열 색상 suffix 제거: CP-XXXNNN{COLOR} → CP-XXXNNN
    """
    d = clean(d_raw)
    if d and d not in ('-',):
        return d
    b = b_first_line.strip()
    if not b:
        return ''
    # 모델코드 패턴: 2글자 대문자 + '-' + 대문자+숫자 조합에서 첫 번째 블록 추출
    m = re.match(r'^([A-Z]{2,}-[A-Z]+\d+)', b)
    return m.group(1) if m else b


def parse_management(i_val):
    """I열 → (managementType, visitCycle)"""
    s = clean(i_val)
    if s == '12개월':
        return '셀프관리', ''
    if not s or '없음' in s:
        return '관리없음', ''
    return '방문관리', s


def parse_commission(k_val, l_val):
    """K/L 수수료 → 유효한 값 중 최댓값 (소수점 그대로, 반올림 없음)"""
    k = to_float(k_val)
    l = to_float(l_val)
    if k is not None and l is not None:
        return max(k, l)
    if k is not None:
        return k
    if l is not None:
        return l
    return 0.0


def is_model_code_pattern(s):
    """B열 값이 모델코드 패턴인지 (한글 없고 영숫자+하이픈만)"""
    return bool(re.match(r'^[A-Z0-9\-\(\)/ ]+$', s.upper()))


def classify_g(g_s):
    """G열 값 분류: 'tasa' / 'promo' / 'blank' / 'other'"""
    if not g_s:
        return 'blank'
    if '타사보상' in g_s:
        return 'tasa'
    if '반값' in g_s:
        return 'promo'
    return 'other'


# ─────────────────────────────────────────────
# 파서
# ─────────────────────────────────────────────
def parse_cuckoo(filepath=TL_PATH, sheet_name=SHEET):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name]

    # ─── 1차 패스: 전체 행 수집 (캐리포워드 + G 분류) ───────────────
    _last_b_first = ''
    _last_a_first = ''
    _last_b_name  = ''

    all_rows = []  # raw row dicts (미필터)

    for r in range(DATA_ROW, ws.max_row + 1):
        a_raw = ws.cell(r, 1).value
        b_raw = ws.cell(r, 2).value
        d_raw = ws.cell(r, 4).value
        e_val = ws.cell(r, 5).value
        g_val = ws.cell(r, 7).value
        h_val = ws.cell(r, 8).value
        i_val = ws.cell(r, 9).value
        j_val = ws.cell(r, 10).value
        k_val = ws.cell(r, 11).value
        l_val = ws.cell(r, 12).value

        # ── A/B 캐리포워드 ──
        a_first = clean(a_raw).split('\n')[0].strip()
        if a_first:
            _last_a_first = a_first

        b_first = clean(b_raw).split('\n')[0].strip()
        if b_first:
            _last_b_first = b_first
            if not is_model_code_pattern(b_first):
                _last_b_name = re.sub(r'^[●○▶▷\*\s]+', '', b_first).strip()
            else:
                _last_b_name = ''

        # 렌탈료 없으면 수집 자체를 스킵
        j = to_float(j_val)
        if not j:
            continue

        # 모델코드
        model_code = normalize_model_code(_last_b_first, d_raw)
        if not model_code:
            continue

        # 약정
        months = parse_months(e_val)
        if months == 0:
            continue

        all_rows.append({
            'modelCode':  model_code,
            'a_first':    _last_a_first,
            'b_name':     _last_b_name,
            'g_class':    classify_g(clean(g_val)),
            'h_val':      clean(h_val),
            'i_val':      i_val,
            'j':          j,
            'k_val':      k_val,
            'l_val':      l_val,
            'months':     months,
        })

    # ─── 프로모션 보유 모델 집합 ──────────────────────────────────────
    # 타사보상 or 반값 행이 하나라도 있는 모델코드
    models_with_promo = {
        row['modelCode']
        for row in all_rows
        if row['g_class'] in ('tasa', 'promo')
    }

    # ─── 2차 패스: G열 필터 적용 후 옵션 생성 ────────────────────────
    raw_options = []

    for row in all_rows:
        g_class = row['g_class']
        mc      = row['modelCode']

        # 채택 규칙:
        #   tasa / promo → 항상 채택
        #   blank → 해당 모델에 promo가 없을 때만 채택
        #   other → 항상 제외
        if g_class == 'other':
            continue
        if g_class == 'blank' and mc in models_with_promo:
            continue

        months   = row['months']
        years    = months // 12
        h_s      = row['h_val']
        is_package = h_s in ('패키지', '패키지10%')
        has_tasa   = (g_class == 'tasa')
        is_promo   = (g_class == 'promo')

        mgmt_type, visit_cycle = parse_management(row['i_val'])
        commission = parse_commission(row['k_val'], row['l_val'])

        name     = row['b_name'] if row['b_name'] else mc
        category = row['a_first']

        raw_options.append({
            'modelCode':      mc,
            'name':           name,
            'category':       category,
            'contractMonths': months,
            'contractYears':  years,
            'contractLabel':  months_to_label(months),
            'managementType': mgmt_type,
            'visitCycle':     visit_cycle,
            'isPackage':      is_package,
            'packageType':    h_s,          # '일반' / '패키지' / '패키지10%'
            'hasTasa':        has_tasa,
            'isPromo':        is_promo,
            'monthlyFee':     row['j'],
            'commission':     commission,
        })

    # ─────────────────────────────────────────────
    # 제품 목록 구성
    # ─────────────────────────────────────────────
    products_map = {}  # modelCode → product dict

    for opt in raw_options:
        mc = opt['modelCode']
        if mc not in products_map:
            products_map[mc] = {
                'id':        mc,
                'modelCode': mc,
                'name':      opt['name'],
                'category':  opt['category'],
                'options':   [],
            }
        else:
            # 제품명 갱신 (한글 제품명 우선)
            if opt['name'] != mc and products_map[mc]['name'] == mc:
                products_map[mc]['name'] = opt['name']

        # ── 패키지 중복 제거 ──
        # 같은 (months, mgmt, visitCycle, hasTasa, isPromo) 조합의 패키지가
        # 여러 개(패키지 / 패키지10%)면 monthlyFee 더 낮은 것만 유지
        if opt['isPackage']:
            pkg_key = (
                opt['contractMonths'],
                opt['managementType'],
                opt['visitCycle'],
                opt['hasTasa'],
                opt['isPromo'],
            )
            existing = next(
                (o for o in products_map[mc]['options']
                 if o['isPackage'] and (
                     o['contractMonths'], o['managementType'], o['visitCycle'],
                     o['hasTasa'], o['isPromo']
                 ) == pkg_key),
                None
            )
            if existing:
                if opt['monthlyFee'] < existing['monthlyFee']:
                    products_map[mc]['options'].remove(existing)
                    products_map[mc]['options'].append(opt)
                # 더 비싸거나 같으면 스킵
                continue

        products_map[mc]['options'].append(opt)

    products = list(products_map.values())

    # ── 통계 ──
    total_opts = sum(len(p['options']) for p in products)
    tasa_opts  = sum(1 for p in products for o in p['options'] if o['hasTasa'])
    promo_opts = sum(1 for p in products for o in p['options'] if o['isPromo'])
    pkg_opts   = sum(1 for p in products for o in p['options'] if o['isPackage'])
    blank_only = len([mc for mc in set(r['modelCode'] for r in all_rows)
                      if mc not in models_with_promo])

    msg = f'[쿠쿠] 파싱 완료: {len(products)}개 제품, {total_opts}개 옵션'
    print(msg.encode('cp949', errors='replace').decode('cp949'))
    msg2 = (f'       타사보상: {tasa_opts}건, 반값: {promo_opts}건, 패키지: {pkg_opts}건\n'
            f'       프로모션 없는 제품(빈칸G 채택): {blank_only}개')
    print(msg2.encode('cp949', errors='replace').decode('cp949'))

    return {
        'metadata': {
            'source':     '쿠쿠',
            'sheetName':  sheet_name,
            'sourceFile': os.path.basename(filepath),
            'parsedAt':   datetime.now().strftime('%Y-%m-%d %H:%M'),
        },
        'products': products,
    }


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────
if __name__ == '__main__':
    import sys
    filepath = sys.argv[1] if len(sys.argv) > 1 else TL_PATH

    data = parse_cuckoo(filepath)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(base_dir, 'cuckoo_data.json')
    with open(out, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    msg = f'저장: {out}'
    print(msg.encode('cp949', errors='replace').decode('cp949'))
    print(f'제품 수: {len(data["products"])}')

    # 샘플 출력
    for p in data['products'][:5]:
        name = p['name']
        mc = p['modelCode']
        print(f'  [{mc}] {name}'.encode('cp949', errors='replace').decode('cp949'))
        for o in p['options'][:4]:
            line = (f'    {o["contractLabel"]} {o["managementType"]} '
                    f'{"패키지("+o["packageType"]+")" if o["isPackage"] else "일반"} '
                    f'{"타사보상" if o["hasTasa"] else ("반값" if o["isPromo"] else "기본")} '
                    f'월{o["monthlyFee"]:,.0f} 수수료{o["commission"]:,.1f}')
            print(line.encode('cp949', errors='replace').decode('cp949'))
