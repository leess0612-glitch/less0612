"""
티엘 SK매직 수수료 엑셀 → JSON 변환 파서
"""
import openpyxl
import json
import re
import os
from datetime import datetime


def clean(v):
    if v is None:
        return ""
    return str(v).strip().replace("\xa0", " ").replace("\u3000", " ")


def normalize_model_code(code):
    return re.sub(r'[\-\s]', '', str(code)).upper()


def parse_tl(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)

    ws = None
    for name in wb.sheetnames:
        if name.strip().upper() == "SK":
            ws = wb[name]
            break
    if ws is None:
        for name in wb.sheetnames:
            if "SK" in name.upper():
                ws = wb[name]
                break
    if ws is None:
        raise ValueError(f"SK 시트를 찾을 수 없습니다. 시트 목록: {wb.sheetnames}")

    rows = list(ws.iter_rows(values_only=True))
    DATA_START = 2  # 행2=헤더, 행3(index 2)부터 데이터

    products = []
    warning_models = set()
    # 접수처 비교용 lookup: key → commission
    # key = "{normalized_model}|{mgmt}|{years}|{tasa}|{package}"
    option_lookup = {}
    # 관리주기 lookup: normalized_model → G열 원본값
    visit_cycle_lookup = {}
    # 모델코드 표기 lookup: normalized_model → TL 원본 표기 (WPU-B600F 형식)
    model_display_map = {}

    current_model_code = ""
    current_product_name = ""

    for i, row in enumerate(rows):
        if i < DATA_START:
            continue

        col_b = clean(row[1])
        col_c = clean(row[2])
        col_e = clean(row[4])
        col_f = clean(row[5])
        col_g = clean(row[6])
        col_h = row[7]
        col_j = row[9]
        col_k = row[10]
        col_l = row[11]

        if col_b:
            current_model_code = col_b
            norm = normalize_model_code(col_b)
            # G열 관리주기 저장 (B열이 있는 행에만 G열 존재, 해당없음 제외)
            if col_g and col_g != '해당없음' and '+' not in col_b:
                visit_cycle_lookup[norm] = col_g
            # 모델코드 원본 표기 저장
            if '+' not in col_b:
                model_display_map[norm] = col_b
        if col_c:
            current_product_name = col_c

        if not current_model_code:
            continue

        # "+" 포함 패키지 모델 제외
        if "+" in current_model_code:
            continue

        # 사업자전용 할인 행 제외 (H열에 '사업자' 포함)
        col_h_s = clean(col_h)
        if '사업자' in col_h_s:
            continue

        # 패키지 감지: "_패키지"(언더스코어) 또는 " 패키지"(공백) 모두 처리
        is_package = "_패키지" in col_f or bool(re.search(r'\s패키지', col_f))

        # 관리구분
        col_e_s = col_e.replace(" ", "")
        if "방문" in col_e_s:
            mgmt = "방문관리"
        elif "셀프" in col_e_s:
            mgmt = "셀프관리"
        elif "관리없음" in col_e_s:
            mgmt = "관리없음"
        else:
            continue

        # 약정년수
        year_match = re.search(r'(\d+)년', col_f)
        if not year_match:
            continue
        contract_years = int(year_match.group(1))
        has_tasa = "_타사보상" in col_f

        # 사이즈 추출 (MAT 제품 F열: 5년약정_K, 5년약정_Q, 5년약정_SS 등)
        size_m = re.search(r'_(K|Q|SS)(?:_|$)', col_f)
        size = size_m.group(1) if size_m else ""

        # 수치
        try:
            monthly_fee = int(col_j) if col_j else 0
            fee_ref     = int(col_k) if col_k else 0
            commission  = int(col_l) if col_l else 0
        except (ValueError, TypeError):
            continue

        if monthly_fee == 0 and commission == 0:
            continue

        # ★ K열은 반값할인가이므로 비교 무의미 → 티엘 dataWarning 없음
        data_warning = False

        # lookup 등록 (패키지/비패키지 모두)
        # 사이즈 있으면 키에 포함 (MAT 사이즈별 수수료 구분)
        norm_code = normalize_model_code(current_model_code)
        if size:
            lookup_key = f"{norm_code}|{mgmt}|{contract_years}|{int(has_tasa)}|{int(is_package)}|{size}"
        else:
            lookup_key = f"{norm_code}|{mgmt}|{contract_years}|{int(has_tasa)}|{int(is_package)}"
        option_lookup[lookup_key] = commission
        # 사이즈 없는 fallback 키도 함께 등록 (사이즈 무관 조회용)
        if size:
            fallback_key = f"{norm_code}|{mgmt}|{contract_years}|{int(has_tasa)}|{int(is_package)}"
            if fallback_key not in option_lookup:
                option_lookup[fallback_key] = commission

        # 제품 목록 (패키지 행은 products에 저장하지 않음 - lookup용으로만 사용)
        if not is_package:
            model_key = normalize_model_code(current_model_code)
            existing = next((p for p in products
                             if normalize_model_code(p["modelCode"]) == model_key), None)
            if existing is None:
                existing = {
                    "modelCode": current_model_code,
                    "name": current_product_name,
                    "options": []
                }
                products.append(existing)

            existing["options"].append({
                "managementType": mgmt,
                "contractYears": contract_years,
                "contractLabel": f"{contract_years}년",
                "hasTasa": has_tasa,
                "size": size,
                "monthlyFee": monthly_fee,
                "commission": commission,
                "dataWarning": data_warning,
            })

    # ── WW / PSG(P↔S) 변형 병합 ──
    # 동일 수수료·요금이면 색상/협업 변형으로 간주 → products 목록에서 제거
    # (optionLookup 항목은 유지 — AK에서 해당 코드로 조회 가능하도록)
    merged_variants = []   # [{"removed": code, "mergedInto": code, "reason": str}]

    def _opts_equal(opts_a, opts_b):
        """옵션 리스트 수수료·요금 비교 (순서 무관)"""
        if len(opts_a) != len(opts_b):
            return False
        key_fn = lambda o: (o["managementType"], o["contractYears"], o.get("hasTasa",False), o.get("size",""))
        a_map = {key_fn(o): (o["monthlyFee"], o["commission"]) for o in opts_a}
        b_map = {key_fn(o): (o["monthlyFee"], o["commission"]) for o in opts_b}
        return a_map == b_map

    # 1) WW 변형 감지 : "(WW)" 괄호형 또는 모델코드 끝이 WW
    to_remove = []
    for p in products:
        mc = p["modelCode"]
        norm_mc = normalize_model_code(mc)
        if "(WW)" in mc:
            base_mc = mc.replace("(WW)", "").strip()
        elif norm_mc.endswith("WW") and len(norm_mc) > 4:
            # 끝 WW 제거 후 base 탐색 (짧은 코드부터 매칭)
            base_mc = None
            for cand in sorted([p2["modelCode"] for p2 in products], key=len):
                cand_norm = normalize_model_code(cand)
                if cand_norm != norm_mc and norm_mc.startswith(cand_norm) and not normalize_model_code(cand).endswith("WW"):
                    base_mc = cand
                    break
            if base_mc is None:
                continue
        else:
            continue

        base = next((x for x in products if x["modelCode"] == base_mc), None)
        if base is None:
            # base_mc에 정확히 일치하는 제품이 없으면 prefix 탐색
            base = next((x for x in products
                         if normalize_model_code(mc).startswith(normalize_model_code(x["modelCode"]))
                         and not normalize_model_code(x["modelCode"]).endswith("WW")
                         and x is not p), None)
        if base and _opts_equal(p["options"], base["options"]):
            to_remove.append(p)
            # lookup 별칭 등록: WW 코드도 base 코드와 동일 커미션으로 이미 등록돼 있으므로 추가 작업 불필요
            merged_variants.append({
                "removed": mc, "mergedInto": base["modelCode"],
                "reason": "위글위글(WW) 색상 변형 — 수수료 동일"
            })
            msg = f"  [WW병합] {mc} → {base['modelCode']}"
            print(msg.encode('cp949', errors='replace').decode('cp949'))

    for p in to_remove:
        products.remove(p)

    # 2) PSG(P 접미사) 변형 감지 : P↔S 쌍, 동일 수수료이면 P 제거
    to_remove_psg = []
    for p in products:
        mc = p["modelCode"]
        norm_mc = normalize_model_code(mc)
        if not norm_mc.endswith("P"):
            continue
        # 같은 prefix + S 모델 탐색
        s_mc_norm = norm_mc[:-1] + "S"
        base_s = next((x for x in products if normalize_model_code(x["modelCode"]) == s_mc_norm), None)
        if base_s and _opts_equal(p["options"], base_s["options"]):
            to_remove_psg.append(p)
            # lookup 별칭: P코드 → S코드 값으로 이미 각각 등록돼 있음
            merged_variants.append({
                "removed": mc, "mergedInto": base_s["modelCode"],
                "reason": "PSG 협업 변형 — 수수료 동일"
            })
            msg = f"  [PSG병합] {mc} → {base_s['modelCode']}"
            print(msg.encode('cp949', errors='replace').decode('cp949'))

    for p in to_remove_psg:
        products.remove(p)

    warn_count = len(warning_models)
    print(f"[티엘] 파싱 완료: {len(products)}개 제품 (병합 {len(merged_variants)}개), "
          f"J≠K 경고 {warn_count}개: {sorted(warning_models) if warn_count else '없음'}")

    return {
        "metadata": {
            "source": "티엘",
            "sheetName": ws.title,
            "sourceFile": os.path.basename(filepath),
            "parsedAt": datetime.now().strftime("%Y-%m-%d %H:%M")
        },
        "products": products,
        "warningModels": sorted(warning_models),
        "optionLookup": option_lookup,
        "visitCycleLookup": visit_cycle_lookup,
        "modelDisplayMap": model_display_map,
        "mergedVariants": merged_variants,
    }


if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        filepath = r"C:\Users\a\Documents\렌탈정책\26.04\2026.04.06 수수료.xlsx"
    else:
        filepath = sys.argv[1]

    data = parse_tl(filepath)

    base_dir = os.path.dirname(os.path.abspath(__file__))
    out = os.path.join(base_dir, "tl_data.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"저장: {out}")
    print(f"lookup 항목수: {len(data['optionLookup'])}")
