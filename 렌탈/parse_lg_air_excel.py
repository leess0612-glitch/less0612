"""
LG 공청기·제습기 수수료 엑셀 → JSON 변환 파서
에이컴즈: ★LG 26년 04월 구독전문점 수수료안내_0401.xlsx (시트: 5)공기청정기)
티엘:     2026.04.09 수수료.xlsx (시트: LG청정+제습)

사용법: python parse_lg_air_excel.py
"""
import openpyxl
import json
import re
import os
import glob as glob_mod
from datetime import datetime


# ─────────────────────────────────────────────
# 경로 / 상수
# ─────────────────────────────────────────────

AC_PATH  = r'C:\Users\a\Documents\렌탈정책\26.04\★LG 26년 04월 구독전문점 수수료안내_0401.xlsx'
TL_PATH  = r'C:\Users\a\Documents\렌탈정책\26.04\2026.04.09 수수료.xlsx'
AC_SHEET = '5)공기청정기'
TL_SHEET = 'LG청정+제습'
DATA_ROW = 5   # 1-indexed 데이터 시작 행

# 렌탈주관사(LG전자) 별도 수수료 — 매달 확인 후 수정
EXTRA_FEE_AK = 100000
EXTRA_FEE_TL = 100000

# 단종 제외 키워드
EXCLUDE_STATUS = {'임시 단종'}

# 가습기 병기 대상 모델 prefix
HUMIDIFIER_PREFIX = 'HY'

# 서비스타입 토글 설명
SERVICE_TYPE_DESC = {
    '라이트플러스': '기본 세척 케어서비스, 제품 성능 점검, 필터 클리닝 및 교체, 무상 A/S 제공',
    '프리미엄':    '분해 세척 케어서비스, 제품 성능 점검, 필터 클리닝 및 교체, 무상 A/S 제공',
    '기본케어':    '방문관리 없음',
}

# 팝업 예외: {정규화모델코드: 메시지}
# 티엘에만 있는 제품 → 에이컴즈 접수 불가
POPUP_TL_ONLY = {
    'AS336NSLCM': '에이컴즈 접수 불가 — 티엘로만 접수 가능\n확인 필요',
    'FS065PSJCM': '에이컴즈 접수 불가 — 티엘로만 접수 가능\n확인 필요',
}
# 에이컴즈에만 있는 제품 → 티엘 접수 불가 (필요 시 추가)
POPUP_AK_ONLY = {
    # 예: 'MODELCODE': '티엘 접수 불가 — 에이컴즈로만 접수 가능\n확인 필요',
}


# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────

def clean(v):
    if v is None:
        return ''
    return str(v).strip().replace('\xa0', ' ').replace('\u3000', ' ')


def to_int(v):
    try:
        return int(v) if v not in (None, '', '-') else 0
    except (ValueError, TypeError):
        return 0


def normalize_model_code(raw):
    """모델코드에서 .AKOR / .AKOR2 / .AKORR 등 suffix 제거."""
    s = clean(raw)
    if not s:
        return ''
    return s.split('.')[0].strip()


def get_lineup_display(lineup, model_code):
    """하이드로타워(HY*)는 (가습기) 병기."""
    if model_code.startswith(HUMIDIFIER_PREFIX):
        if '(가습기)' not in lineup:
            return lineup + ' (가습기)'
    return lineup


# ─────────────────────────────────────────────
# 헤더 스캔 — 동적 열 매핑
# ─────────────────────────────────────────────

def scan_columns(ws):
    """
    헤더 행(1~4)을 읽어 월요금 열과 수수료 열을 동적으로 매핑.

    반환:
        fee_cols  : {(contractMonths, combineType): col_idx}  ← 월요금
        comm_cols : {(contractMonths, combineType): col_idx}  ← 수수료
        warnings  : [str]  ← 예상치 못한 헤더 변경 경고
    """
    max_col = ws.max_column

    def cell_val(r, c):
        return ws.cell(r, c).value

    # 기대하는 (months, combine) 조합
    EXPECTED_KEYS = {
        (36, '단품'), (48, '단품'), (48, '신규결합'),
        (60, '단품'), (60, '신규결합'),
        (72, '단품'), (72, '신규결합'), (72, '기존결합'),
    }

    fee_cols  = {}
    comm_cols = {}
    current_month = None

    for c in range(1, max_col + 1):
        r3 = clean(cell_val(3, c))
        r4 = clean(cell_val(4, c))

        # 월 갱신 (row 3에 "N개월" 형태)
        m = re.match(r'^(\d+)개월$', r3)
        if m:
            current_month = int(m.group(1))
        elif r3 and r3 not in ('', '구독료', '수수료'):
            # 월 관련 없는 레이블이면 current_month 리셋
            if not any(kw in r3 for kw in ('개월', '수수료', '선납', '구독료')):
                current_month = None

        if current_month is None:
            continue

        # row 4 기준 결합타입 분류
        r4_norm = r4.replace('\n', ' ')

        # 월요금 열
        if r4_norm == '구독료':
            fee_cols[(current_month, '단품')] = c
        elif '신규결합' in r4_norm and '시' not in r4_norm:
            fee_cols[(current_month, '신규결합')] = c
        elif '기존결합' in r4_norm and '시' not in r4_norm:
            fee_cols[(current_month, '기존결합')] = c
        # 수수료 열
        elif r4_norm == '수수료':
            comm_cols[(current_month, '단품')] = c
        elif r4_norm == '신규결합시':
            comm_cols[(current_month, '신규결합')] = c
        elif r4_norm == '기존결합시':
            comm_cols[(current_month, '기존결합')] = c

    # 경고 생성
    warnings = []
    found_fee_keys  = set(fee_cols.keys())
    found_comm_keys = set(comm_cols.keys())

    for key in EXPECTED_KEYS:
        if key not in found_fee_keys:
            warnings.append(f'⚠️  월요금 열 누락: {key[0]}개월 {key[1]}')
        if key not in found_comm_keys:
            warnings.append(f'⚠️  수수료 열 누락: {key[0]}개월 {key[1]}')

    for key in found_fee_keys - EXPECTED_KEYS:
        warnings.append(f'ℹ️  새 월요금 열 발견: {key[0]}개월 {key[1]} — 파서 추가 여부 확인')
    for key in found_comm_keys - EXPECTED_KEYS:
        warnings.append(f'ℹ️  새 수수료 열 발견: {key[0]}개월 {key[1]} — 파서 추가 여부 확인')

    return fee_cols, comm_cols, warnings


# ─────────────────────────────────────────────
# 시트 파싱
# ─────────────────────────────────────────────

def parse_sheet(wb, sheet_name, extra_fee, fee_cols, comm_cols):
    """
    시트를 읽어 {(model_code, visit_cycle, service_type): row_data} 반환.
    row_data = {
        'lineup': str,
        'productType': str,
        'visitCycle': int or '자가관리',
        'serviceType': str,
        'fees': {(contractMonths, combineType): monthlyFee},
        'commissions': {(contractMonths, combineType): commission + extra_fee},
    }
    """
    ws = wb[sheet_name]
    result = {}

    for r in range(DATA_ROW, ws.max_row + 1):
        def cv(c):
            return ws.cell(r, c).value

        status       = clean(cv(2))   # B열: 전월대비
        product_type = clean(cv(4))   # D열: 제품군
        lineup_raw   = clean(cv(5))   # E열: 라인업
        model_raw    = clean(cv(6))   # F열: 모델코드
        visit_raw    = cv(7)          # G열: 방문주기
        service_type = clean(cv(8))   # H열: 서비스타입

        # 단종 제외
        if status in EXCLUDE_STATUS:
            continue
        # 빈 행 스킵
        if not model_raw or not lineup_raw:
            continue

        model_code = normalize_model_code(model_raw)
        lineup     = get_lineup_display(lineup_raw, model_code)

        # 방문주기 정규화
        if clean(str(visit_raw)) == '자가관리':
            visit_cycle = '자가관리'
        else:
            try:
                visit_cycle = int(visit_raw)
            except (TypeError, ValueError):
                continue

        # 월요금 / 수수료 읽기 (값 있는 것만)
        fees = {}
        for key, col in fee_cols.items():
            val = to_int(cv(col))
            if val > 0:
                fees[key] = val

        commissions = {}
        for key, col in comm_cols.items():
            val = to_int(cv(col))
            if val > 0:
                commissions[key] = val + extra_fee

        if not fees and not commissions:
            continue

        row_key = (model_code, visit_cycle, service_type)
        result[row_key] = {
            'lineup':      lineup,
            'productType': product_type,
            'visitCycle':  visit_cycle,
            'serviceType': service_type,
            'fees':        fees,
            'commissions': commissions,
        }

    return result


# ─────────────────────────────────────────────
# 제품 목록 생성 (에이컴즈 + 티엘 비교)
# ─────────────────────────────────────────────

def build_products(ac_data, tl_data):
    all_keys = set(ac_data.keys()) | set(tl_data.keys())

    # 모델코드별로 그룹화
    model_groups = {}
    for key in all_keys:
        mc = key[0]
        model_groups.setdefault(mc, set()).add(key)

    products = []

    for model_code in sorted(model_groups):
        keys = sorted(model_groups[model_code])
        sample = (ac_data.get(keys[0]) or tl_data.get(keys[0], {}))

        options = []

        for row_key in keys:
            _, visit_cycle, service_type = row_key
            ac_row = ac_data.get(row_key)
            tl_row = tl_data.get(row_key)

            # 모든 (contractMonths, combineType) 조합 수집
            all_combo_keys = set()
            if ac_row:
                all_combo_keys |= set(ac_row['fees'].keys())
            if tl_row:
                all_combo_keys |= set(tl_row['fees'].keys())

            for combo_key in sorted(all_combo_keys):
                contract_months, combine_type = combo_key

                ac_fee_val = (ac_row['commissions'].get(combo_key, 0)) if ac_row else 0
                tl_fee_val = (tl_row['commissions'].get(combo_key, 0)) if tl_row else 0
                monthly_fee = (
                    (ac_row['fees'].get(combo_key) if ac_row else None)
                    or (tl_row['fees'].get(combo_key) if tl_row else None)
                    or 0
                )

                # 접수처 결정
                if ac_fee_val > 0 and tl_fee_val > 0:
                    if ac_fee_val > tl_fee_val:
                        recommended = '에이컴즈'
                    elif tl_fee_val > ac_fee_val:
                        recommended = '티엘'
                    else:
                        recommended = '동일'
                elif ac_fee_val > 0:
                    recommended = '에이컴즈'
                elif tl_fee_val > 0:
                    recommended = '티엘'
                else:
                    recommended = None

                # 팝업 결정
                popup = None
                if model_code in POPUP_TL_ONLY and not ac_row:
                    popup = POPUP_TL_ONLY[model_code]
                elif model_code in POPUP_AK_ONLY and not tl_row:
                    popup = POPUP_AK_ONLY[model_code]

                options.append({
                    'visitCycle':        visit_cycle,
                    'serviceType':       service_type,
                    'serviceDesc':       SERVICE_TYPE_DESC.get(service_type),
                    'contractMonths':    contract_months,
                    'combineType':       combine_type,
                    'monthlyFee':        monthly_fee,
                    'commission': {
                        'ak': ac_fee_val or None,
                        'tl': tl_fee_val or None,
                    },
                    'recommendedOffice': recommended,
                    'popup':             popup,
                })

        products.append({
            'id':          model_code,
            'modelCode':   model_code,
            'lineup':      sample.get('lineup', model_code),
            'productType': sample.get('productType', '공기청정기'),
            'options':     options,
        })

    return products


# ─────────────────────────────────────────────
# HTML 주입
# ─────────────────────────────────────────────

def inject_into_html(data, base_dir, placeholder='__LG_AIR_DATA__'):
    pattern = os.path.join(base_dir, '렌탈수수료_[0-9]*.html')
    candidates = sorted(glob_mod.glob(pattern))
    if not candidates:
        print('⚠️  렌탈수수료_*.html 파일을 찾지 못했습니다. parse_excel.py를 먼저 실행하세요.')
        return
    target = candidates[-1]
    with open(target, 'r', encoding='utf-8') as f:
        html = f.read()
    if placeholder not in html:
        print(f'⚠️  {os.path.basename(target)} 에 {placeholder} 플레이스홀더가 없습니다.')
        return
    html_out = html.replace(placeholder, json.dumps(data, ensure_ascii=False))
    with open(target, 'w', encoding='utf-8') as f:
        f.write(html_out)
    print(f'HTML 주입 완료: {target}')


# ─────────────────────────────────────────────
# 메인
# ─────────────────────────────────────────────

def parse_lg_air(ac_filepath=AC_PATH, tl_filepath=TL_PATH):
    base_dir = os.path.dirname(os.path.abspath(__file__))

    print(f'[에이컴즈] 파싱 중: {os.path.basename(ac_filepath)}')
    ac_wb = openpyxl.load_workbook(ac_filepath, data_only=True)
    ac_ws = ac_wb[AC_SHEET]
    ac_fee_cols, ac_comm_cols, ac_warnings = scan_columns(ac_ws)

    print(f'[티엘]     파싱 중: {os.path.basename(tl_filepath)}')
    tl_wb = openpyxl.load_workbook(tl_filepath, data_only=True)
    tl_ws = tl_wb[TL_SHEET]
    tl_fee_cols, tl_comm_cols, tl_warnings = scan_columns(tl_ws)

    # 경고 출력
    if ac_warnings:
        print('\n[에이컴즈 헤더 경고]')
        for w in ac_warnings:
            print(' ', w)
    if tl_warnings:
        print('\n[티엘 헤더 경고]')
        for w in tl_warnings:
            print(' ', w)

    ac_data = parse_sheet(ac_wb, AC_SHEET, EXTRA_FEE_AK, ac_fee_cols, ac_comm_cols)
    tl_data = parse_sheet(tl_wb, TL_SHEET, EXTRA_FEE_TL, tl_fee_cols, tl_comm_cols)

    products = build_products(ac_data, tl_data)

    # 통계
    total_opts = sum(len(p['options']) for p in products)
    ak_only    = sum(1 for p in products for o in p['options']
                     if o['commission']['ak'] and not o['commission']['tl'])
    tl_only    = sum(1 for p in products for o in p['options']
                     if o['commission']['tl'] and not o['commission']['ak'])
    diff_count = sum(1 for p in products for o in p['options']
                     if o['recommendedOffice'] not in ('동일', None))
    popup_count = sum(1 for p in products for o in p['options'] if o.get('popup'))

    print(f'\n에이컴즈 행수:  {len(ac_data)}개')
    print(f'티엘     행수:  {len(tl_data)}개')
    print(f'제품 수:        {len(products)}개')
    print(f'총 옵션 수:     {total_opts}개')
    print(f'에이컴즈 전용: {ak_only}건')
    print(f'티엘 전용:     {tl_only}건')
    print(f'수수료 차이:   {diff_count}건')
    print(f'팝업 예외:     {popup_count}건')

    data = {
        'metadata': {
            'brand':    'LG전자',
            'category': '공청기·제습기',
            'parsedAt': datetime.now().strftime('%Y-%m-%d'),
            'extraFee': {'ak': EXTRA_FEE_AK, 'tl': EXTRA_FEE_TL},
        },
        'products': products,
    }

    # JSON 저장
    out_path = os.path.join(base_dir, 'lg_air_data.json')
    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f'\nJSON 저장: {out_path}')

    # HTML 주입
    inject_into_html(data, base_dir)

    return data


if __name__ == '__main__':
    parse_lg_air()
