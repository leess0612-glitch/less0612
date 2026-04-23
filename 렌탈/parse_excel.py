"""
SK매직 수수료 엑셀 → JSON 변환 파서
사용법: python parse_excel.py "파일경로.xlsx"
"""
import openpyxl
import json
import sys
import re
import os
from datetime import datetime

# ─────────────────────────────────────────────
# 유틸
# ─────────────────────────────────────────────

def clean(v):
    if v is None:
        return ""
    return str(v).strip().replace("\xa0", " ").replace("\u3000", " ")

def clean_name(v):
    s = clean(v)
    s = re.sub(r'\n+', '\n', s)
    return s.strip()

def months_to_label(m):
    m = int(m)
    if m % 12 == 0:
        return f"{m//12}년"
    return f"{m}개월"

def normalize_management_type(raw):
    if not raw:
        return None
    s = clean(raw).replace("\n", "").replace(" ", "")
    patterns = [
        ("방문할인+타사보상", "방문할인+타사보상"),
        ("셀프할인+타사보상", "셀프할인+타사보상"),
        ("방문할인", "방문할인"),
        ("셀프할인", "셀프할인"),
        ("무방문형할인", "무방문형할인"),
        ("방문", "방문"),
        ("셀프", "셀프"),
        ("무방문형", "무방문형"),
        ("타사보상", "타사보상"),
    ]
    year_mgmt = re.match(r'^(\d+)년\(?(방문형|셀프형|무방문형)\)?$', s)
    if year_mgmt:
        return s
    for keyword, label in patterns:
        if keyword in s:
            return label
    return None

def detect_category(model_code, product_name, row_index):
    m = (model_code + product_name).upper()
    if any(x in m for x in ["WPU", "정수기", "언더싱크", "그랜드정수기", "뉴랜드정수기", "뉴슬림정수기"]):
        return "정수기"
    if any(x in m for x in ["ACL", "공기청정기", "청정기"]):
        return "공기청정기"
    if any(x in m for x in ["BID", "비데"]):
        return "비데"
    if any(x in m for x in ["MAT", "매트리스", "워커힐", "에코휴", "파운데이션", "헤드보드", "PVC", "레더"]):
        return "매트리스"
    if any(x in m for x in ["구독", "선결제", "일시불", "멤버쉽"]):
        return "기타"
    return "기타"

def parse_product_name_from_col2(raw):
    if not raw:
        return "", ""
    parts = [p.strip() for p in str(raw).replace("\xa0"," ").split("\n") if p.strip()]
    if not parts:
        return "", ""
    first = parts[0]
    model_code = ""
    name_parts = []
    if re.match(r'^[A-Z0-9\-]+$', first.replace(" ","").upper()) and len(first) < 30:
        model_code = first.strip()
        name_parts = parts[1:]
    else:
        name_parts = parts
    name = " ".join(name_parts).strip()
    return model_code, name

def clean_option_name(col4_raw, model_code):
    s = clean(col4_raw)
    lite = False
    if s.startswith("라이트시리즈"):
        lite = True
        s = s[len("라이트시리즈"):].strip()

    discounts = re.findall(r'\([\d,]+원\s*할인\)', s)

    if model_code and s.upper().replace(" ","").startswith(model_code.upper().replace(" ","")):
        s = s[len(model_code):].strip()

    s = re.sub(r'^[A-Z0-9,–\-\s]+(?=\(|$)', '', s).strip()
    s = re.sub(r'\(방문\)', '', s)
    s = re.sub(r'\(셀프\)', '', s)
    s = re.sub(r'\(\d+년의무\)', '', s)
    s = re.sub(r'\(\d+년\)', '', s)
    s = re.sub(r'[A-Z0-9]+(SK|ASK|CSK)[A-Z0-9]+(,[A-Z0-9]+(SK|ASK|CSK)[A-Z0-9]+)*', '', s)
    s = s.strip(" ,()-+")

    if discounts and not any(re.search(r'[\d,]+원\s*할인', s) for _ in [1]):
        d_clean = discounts[0].strip("()").replace("  "," ")
        s = (s + " " + d_clean).strip()

    if lite:
        s = ("라이트" + (" " + s if s else "")).strip()

    return s if s else ""

# ─────────────────────────────────────────────
# 단종 모델 목록
# ─────────────────────────────────────────────
DISCONTINUED_MODELS = set()  # 현재 단종 모델 없음 (KM720R/QM720R은 TL 미취급이나 단종 아님)

# D열 기본 관리유형 (할인 없음) → 제외 대상
BASIC_MGMT_EXCLUDE = {'방문', '셀프', '무방문형'}

# ─────────────────────────────────────────────
# 메인 파서
# ─────────────────────────────────────────────

def parse_excel(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]
    sheet_title = ws.title.strip()
    rows = list(ws.iter_rows(values_only=True))

    DATA_START = 7

    # ── 1차 스캔: 모든 모델별 할인 옵션 존재 여부 확인 ──
    # 할인 옵션이 있는 모델 → 할인 옵션만 유효
    # 할인 옵션이 없는 모델 → 기본 옵션("방문"/"셀프")도 유효 데이터
    model_has_discount = {}  # model_code(upper) → bool
    scan_model = None
    scan_is_mat = False

    for row in rows[DATA_START:]:
        col2 = row[2]
        col3 = row[3]   # D열: 관리유형 (비MAT 할인 감지)
        col4 = row[4]   # E열: 옵션명   (MAT 할인 감지)

        if col2 is not None and str(col2).strip():
            model_code, _ = parse_product_name_from_col2(col2)
            scan_model = model_code.upper() if model_code else None
            # C열에 모델코드가 없으면 같은 행의 E열에서 추출
            # * 또는 / 있으면 제외 — 복수 코드는 별도 분리 로직(_no_model_e_map)이 처리
            if not scan_model and col4 and '*' not in str(col4) and '/' not in str(col4):
                _e_codes = extract_e_model_codes(clean(col4))
                if _e_codes:
                    scan_model = _e_codes[0]
            scan_is_mat = bool(scan_model and scan_model.startswith('MAT'))
            if scan_model and scan_model not in model_has_discount:
                model_has_discount[scan_model] = False

        if scan_model:
            if scan_is_mat and col4 and '할인' in str(col4):
                model_has_discount[scan_model] = True
            elif not scan_is_mat and col3 and '할인' in str(col3):
                model_has_discount[scan_model] = True

    mat_has_discount = {k: v for k, v in model_has_discount.items() if k.startswith('MAT')}
    print(f"MAT 할인 스캔 결과: {sum(1 for v in mat_has_discount.values() if v)}개 모델에 할인 옵션 존재")

    # ── 2차 처리: 메인 파싱 ──
    products = []
    current_product = None
    current_model_code = ""
    current_product_name = ""
    current_category = ""
    current_mgmt_type = ""

    for i, row in enumerate(rows):
        if i < DATA_START:
            continue

        col1 = clean(row[1])
        col2 = row[2]
        col3 = clean(row[3])
        col4 = clean(row[4])
        fee_main   = row[5]   # F열: 실제 월요금 (기준)
        fee_guide  = row[6]   # G열: 가이드 월요금 (참고, 오류감지용)
        obligation = row[8]
        ownership  = clean(row[9])
        reg_fee    = row[10]
        base_comm  = row[11]
        add_cnt    = row[12]
        add_comm   = row[13]
        bonus_comm = row[14]
        total_comm = row[15]

        # 새 제품 그룹 시작
        if col2 is not None and str(col2).strip():
            model_code, product_name = parse_product_name_from_col2(col2)
            # C열에 모델코드 없으면 같은 행 E열에서 추출
            # * 또는 / 있으면 제외 — 복수 코드는 별도 분리 로직(_no_model_e_map)이 처리
            if not model_code and col4 and '*' not in str(col4) and '/' not in str(col4):
                _e_codes = extract_e_model_codes(col4)
                if _e_codes:
                    model_code = _e_codes[0]
            category = detect_category(model_code, product_name, i)

            # 구독/선결제/일시불/멤버쉽 섹션 스킵
            if any(k in (model_code + product_name) for k in ["구독", "선결제", "일시불", "멤버쉽"]):
                current_product = None
                current_model_code = ""
                current_product_name = ""
                current_category = category
                continue

            # ★ 단종 제품 제외
            model_upper = model_code.upper()
            if model_upper in DISCONTINUED_MODELS:
                current_product = None
                current_model_code = ""
                current_product_name = ""
                current_category = category
                print(f"  단종 제외: {model_code}")
                continue

            current_product = {
                "id": (model_code or product_name[:10]).replace(" ",""),
                "modelCode": model_code,
                "name": product_name,
                "category": category,
                "promotionNote": "",
                "note": "",
                "options": []
            }
            current_model_code = model_code
            current_product_name = product_name
            current_category = category
            current_mgmt_type = ""
            products.append(current_product)

        if current_product is None:
            continue

        if col1:
            current_product["promotionNote"] = col1

        # 관리유형 처리
        mgmt_type = normalize_management_type(col3)
        if col3 and mgmt_type is None:
            special_note = clean(col3).replace("\n", " ").strip()
            if special_note and special_note not in current_product.get("note",""):
                current_product["note"] = (current_product.get("note","") + " " + special_note).strip()
        elif col3 and mgmt_type:
            current_mgmt_type = mgmt_type
        elif not col3:
            mgmt_type = current_mgmt_type
        if mgmt_type is None:
            mgmt_type = current_mgmt_type

        # ★ E열 "라이트시리즈" → 비데에서만 셀프관리 강제
        # (공기청정기의 "라이트시리즈"는 제품 시리즈명 — 방문관리)
        if "라이트시리즈" in col4 and current_category == "비데":
            mgmt_type = "셀프관리"
            current_mgmt_type = "셀프관리"
        # ★ 비데/공기청정기/정수기에서 D열 비어있고 mgmt_type이 미결정일 때만 방문관리 기본값
        # mgmt_type이 이미 carry-forward로 설정된 경우(타사보상 등) 덮어쓰지 않음
        elif current_category in ("비데", "공기청정기", "정수기") and not col3 and current_model_code and mgmt_type is None:
            mgmt_type = "방문관리"
            current_mgmt_type = "방문관리"

        # ★ MAT 제품: D열 없음 → H열(visitCycle)로 관리방식 결정
        # H열 '없음' = 관리없음(셀프형), H열 '4개월' 등 = 방문관리
        if current_category == "매트리스" and not col3:
            visit_cycle_raw = clean(row[7]) if row[7] is not None else ""
            if visit_cycle_raw == "없음" or visit_cycle_raw == "":
                mgmt_type = "관리없음"
            else:
                mgmt_type = "방문관리"

        # ★ D열 "방문"/"셀프" 기본 행 제외
        # 단, 해당 모델에 할인 옵션이 없으면 기본 행도 유효 데이터로 포함
        if mgmt_type in BASIC_MGMT_EXCLUDE:
            if model_has_discount.get(current_model_code.upper(), True):
                continue

        # 수치 데이터
        try:
            monthly_fee  = int(fee_main)   if fee_main   else 0   # F열 기준
            monthly_ref  = int(fee_guide)  if fee_guide  else 0   # G열 참고
            months       = int(obligation) if obligation else 0
            base_c       = int(base_comm)  if base_comm  else 0
            add_c        = int(add_comm)   if add_comm   else 0
            bonus_c      = int(bonus_comm) if bonus_comm else 0
            total_c      = int(total_comm) if total_comm else 0
            reg          = int(reg_fee)    if reg_fee    else 0
        except (ValueError, TypeError):
            continue

        if monthly_fee == 0 and total_c == 0:
            continue

        # ★ 요금 오류 감지: F열 ≠ G열이면 경고 플래그
        data_warning = (monthly_fee != 0 and monthly_ref != 0 and monthly_fee != monthly_ref)

        # ★ MAT 모델: 할인 옵션 있는 모델은 할인 옵션만 유효
        if current_category == '매트리스':
            mat_key = current_model_code.upper()
            if mat_key in mat_has_discount and mat_has_discount[mat_key]:
                if '할인' not in col4:
                    continue

        option_label = clean_option_name(col4, current_model_code)

        own_months = 0
        own_match = re.search(r'(\d+)', ownership)
        if own_match:
            own_months = int(own_match.group(1))

        option = {
            "label": option_label,
            "managementType": mgmt_type or "",
            "contractMonths": months,
            "contractLabel": months_to_label(months) if months else "",
            "monthlyFee": monthly_fee,
            "dataWarning": data_warning,
            "visitCycle": clean(row[7]),
            "ownershipMonths": own_months,
            "registrationFee": reg,
            "baseCommission": base_c,
            "additionalCount": int(add_cnt) if add_cnt else 0,
            "additionalCommission": add_c,
            "bonusCommission": bonus_c,
            "totalCommission": total_c,
        }
        current_product["options"].append(option)

    # 빈 옵션 제품 제거
    products = [p for p in products if p["options"]]

    # 중복 id 처리: 같은 모델코드면 options 합치기
    merged = {}
    for p in products:
        pid = p["id"]
        if pid in merged:
            merged[pid]["options"].extend(p["options"])
        else:
            merged[pid] = p
    products = list(merged.values())

    return {
        "metadata": {
            "brand": "SK매직",
            "sheetTitle": sheet_title,
            "sourceFile": os.path.basename(filepath),
            "parsedAt": datetime.now().strftime("%Y-%m-%d %H:%M")
        },
        "products": products
    }


# ─────────────────────────────────────────────
# 접수처 비교 헬퍼
# ─────────────────────────────────────────────

def _norm_model(code):
    return re.sub(r'[\-\s]', '', str(code)).upper()

def _fmt_code_display(norm):
    """정규화된 모델코드를 읽기 쉬운 형식으로 변환 (WPU1234 → WPU-1234)"""
    m = re.match(r'^(WPU|ACL|BID|MAT)(.+)$', norm)
    if m:
        return f"{m.group(1)}-{m.group(2)}"
    return norm

def _tl_lookup_key(model_code, tl_mgmt, years, has_tasa, is_package):
    return f"{_norm_model(model_code)}|{tl_mgmt}|{years}|{int(has_tasa)}|{int(is_package)}"

def _tl_model_variants(code):
    """에이컴즈 모델코드 → 티엘 매칭용 변형 목록

    MAT 사이즈 매핑 패턴:
      MAT[KQS]M[숫자]...  → strip [KQS]M  (워커힐 스위트/스탠다드: SM730→730)
      MAT[KQS][비M문자]... → strip [KQS]   (에코휴/헤드보드: SD011→D011, SH510→H510)
      MAT-TSM...           → MAT-SM...     (오타 수정)
    """
    norm = _norm_model(code)
    variants = [norm]

    # MAT-TSM… → MAT-SM… (에이컴즈 오타 수정)
    mat_t = re.match(r'^MATT([^T].*)$', norm)
    if mat_t:
        variants.append(f"MAT{mat_t.group(1)}")

    # MAT + size(KQS) + M + 숫자... → strip size and M (워커힐 스위트/스탠다드)
    mat_sm = re.match(r'^MAT([KQS])M(\d.*)$', norm)
    if mat_sm:
        stripped = f"MAT{mat_sm.group(2)}"
        if stripped not in variants:
            variants.append(stripped)

    # MAT + size(KQS) + 비M문자... → strip size only (에코휴/헤드보드/파운데이션)
    mat_sl = re.match(r'^MAT([KQS])([A-LN-Z].*)$', norm)  # M 제외
    if mat_sl:
        stripped = f"MAT{mat_sl.group(2)}"
        if stripped not in variants:
            variants.append(stripped)

    return variants


def _get_mat_size(model_code):
    """에이컴즈 MAT 모델코드에서 사이즈 추출 (K/Q/SS)"""
    norm = _norm_model(model_code)
    # en dash 정규화
    norm = norm.replace('\u2013', '').replace('\u2014', '')
    m = re.match(r'^MAT([KQS])', norm)
    if m:
        letter = m.group(1)
        return {'K': 'K', 'Q': 'Q', 'S': 'SS'}[letter]
    return ""

def extract_e_model_codes(e_val):
    """E열 값에서 모델코드 추출.
    - 라이트시리즈 prefix 제거
    - (N년의무/의문) 등 괄호 제거
    - 공백/하이픈 정규화
    - * → C형, F형 두 개 반환
    - 공백/슬래시로 구분된 복수 코드 처리
    """
    if not e_val:
        return []
    # 라이트시리즈 prefix 제거 (셀프관리 표시용 텍스트)
    s = e_val
    if s.startswith("라이트시리즈"):
        s = s[len("라이트시리즈"):].strip()
    # 괄호 앞부분만 사용
    base = re.split(r'\(', s)[0].strip()
    # en dash → 일반 hyphen
    base = base.replace('\u2013', '-').replace('\u2014', '-')
    # 슬래시로 분리된 복수 코드 처리 (MAT-SF520RKIV/MAT-SF530RKBE 등)
    parts = [p.strip() for p in base.split('/') if p.strip()]
    results = []
    for part in parts:
        # 라이트시리즈 제거 (모델코드 뒤에 붙는 경우: "ACL130Z0SKPN 라이트시리즈")
        part = re.sub(r'\s*라이트시리즈\s*', '', part).strip()
        norm = re.sub(r'[\-\s]', '', part).upper()
        if not norm or len(norm) < 4:
            continue
        if '*' in norm:
            results.append(norm.replace('*', 'C'))
            results.append(norm.replace('*', 'F'))
        else:
            results.append(norm)
    return results

def tl_match_model(norm_code, tl_known_models):
    """정규화된 에이컴즈 E열 모델코드를 TL 모델코드로 매핑.
    - 정확히 일치: 그대로 사용
    - TL 코드가 더 긴 경우 (에이컴즈 truncated): TL 코드 사용
    - 에이컴즈 코드가 더 긴 경우 (후행 색상코드): TL 코드로 트리밍
    - MAT 사이즈 제거 후 재시도
    """
    if norm_code in tl_known_models:
        return norm_code
    # TL이 더 긴 경우 (짧은 TL 코드부터 확인)
    for tl in sorted(tl_known_models, key=len):
        if tl.startswith(norm_code) and len(tl) > len(norm_code):
            return tl
    # 에이컴즈가 더 긴 경우 (긴 TL 코드부터 확인하여 가장 긴 prefix 매칭)
    matched = [tl for tl in tl_known_models if norm_code.startswith(tl)]
    if matched:
        return max(matched, key=len)
    # MAT 사이즈 제거 후 재시도
    for stripped in _tl_model_variants(norm_code)[1:]:  # 첫 번째는 원본
        if stripped in tl_known_models:
            return stripped
        for tl in sorted(tl_known_models, key=len):
            if tl.startswith(stripped) and len(tl) > len(stripped):
                return tl
        matched2 = [tl for tl in tl_known_models if stripped.startswith(tl)]
        if matched2:
            return max(matched2, key=len)
    return norm_code

def _extend_model_variants_with_prefix(base_variants, tl_known_models):
    """base_variants 에 prefix 매칭으로 찾은 TL 모델코드도 추가"""
    extended = list(base_variants)
    for variant in base_variants:
        # TL 코드가 더 긴 경우 (AK가 prefix)
        for tl in sorted(tl_known_models, key=len):
            if tl.startswith(variant) and tl not in extended:
                extended.append(tl)
        # AK 코드가 더 긴 경우 (TL이 prefix)
        for tl in sorted(tl_known_models, key=len, reverse=True):
            if variant.startswith(tl) and tl not in extended:
                extended.append(tl)
    return extended


def compute_recommended_office(tl_lookup, model_code, mgmt_type, contract_months,
                                ak_commission, is_package=False, tl_known_models=None):
    """에이컴즈 옵션 하나에 대해 접수처 추천을 반환.

    Returns:
        "에이컴즈" | "티엘" | "동일" | None(TL 매칭 없음)
    """
    # ── 관리방식 → TL 관리방식 매핑 (복수 후보 허용) ──
    # MAT 헤드보드/파운데이션은 AK에서 방문관리로 파싱되더라도
    # TL에서 관리없음으로 등록된 경우가 있어 양쪽 모두 시도
    is_mat = _norm_model(model_code).startswith('MAT')

    if not mgmt_type:
        tl_mgmts = ["방문관리", "셀프관리", "관리없음"]
    elif "방문" in mgmt_type:
        # MAT 제품은 관리없음도 fallback 시도 (헤드보드/파운데이션 AK↔TL 불일치)
        tl_mgmts = ["방문관리", "관리없음"] if is_mat else ["방문관리"]
    elif "셀프" in mgmt_type:
        tl_mgmts = ["셀프관리", "관리없음"]
    elif mgmt_type == "관리없음":
        tl_mgmts = ["관리없음", "셀프관리"]
    elif "무방문형" in mgmt_type:
        tl_mgmts = ["셀프관리", "방문관리"]
    else:
        tl_mgmts = ["방문관리", "셀프관리", "관리없음"]  # fallback: 모두 시도

    years = contract_months // 12
    has_tasa = "타사보상" in (mgmt_type or "")
    size = _get_mat_size(model_code)

    # ── 모델코드 변형 목록 (MAT 사이즈 제거 + prefix 매칭) ──
    base_variants = _tl_model_variants(model_code)
    if tl_known_models:
        model_keys = _extend_model_variants_with_prefix(base_variants, tl_known_models)
    else:
        model_keys = base_variants

    tl_commission = None
    for model_key in model_keys:
        for tl_mgmt in tl_mgmts:
            if size:
                key = f"{model_key}|{tl_mgmt}|{years}|{int(has_tasa)}|{int(is_package)}|{size}"
                if key in tl_lookup:
                    tl_commission = tl_lookup[key]
                    break
            key = f"{model_key}|{tl_mgmt}|{years}|{int(has_tasa)}|{int(is_package)}"
            if key in tl_lookup:
                tl_commission = tl_lookup[key]
                break
        if tl_commission is not None:
            break

    # 패키지 lookup 실패 시 TL 일반 수수료로 fallback 비교
    # (TL 엑셀에 패키지 행이 없으므로 일반 수수료 = TL 패키지 수수료)
    if tl_commission is None and is_package:
        for model_key in model_keys:
            for tl_mgmt in tl_mgmts:
                if size:
                    key = f"{model_key}|{tl_mgmt}|{years}|{int(has_tasa)}|0|{size}"
                    if key in tl_lookup:
                        tl_commission = tl_lookup[key]
                        break
                key = f"{model_key}|{tl_mgmt}|{years}|{int(has_tasa)}|0"
                if key in tl_lookup:
                    tl_commission = tl_lookup[key]
                    break
            if tl_commission is not None:
                break

    if tl_commission is None:
        return {'office': None, 'tlCommission': None}

    if ak_commission > tl_commission:
        return {'office': '에이컴즈', 'tlCommission': tl_commission}
    elif tl_commission > ak_commission:
        return {'office': '티엘', 'tlCommission': tl_commission}
    else:
        return {'office': '동일', 'tlCommission': tl_commission}


if __name__ == "__main__":
    if len(sys.argv) < 2:
        filepath_sk = r"C:\Users\a\Documents\렌탈정책\26.04\SK 수수료표_2604v1 (1).xlsx"
        filepath_tl = r"C:\Users\a\Documents\렌탈정책\26.04\2026.04.21 수수료.xlsx"
    else:
        filepath_sk = sys.argv[1]
        filepath_tl = sys.argv[2] if len(sys.argv) > 2 else r"C:\Users\a\Documents\렌탈정책\26.04\2026.04.06 수수료.xlsx"

    base_dir = os.path.dirname(os.path.abspath(__file__))

    # ── 티엘 파싱 (에이컴즈보다 먼저 — 모델코드 정규화에 TL 사용) ──
    tl_warning_models = []
    tl_lookup = {}
    tl_visit_cycle = {}
    tl_model_display = {}
    tl_known_models = set()
    tl_data = {"products": []}
    try:
        from parse_tl_excel import parse_tl
        print(f"[티엘] 파싱 중: {filepath_tl}")
        tl_data = parse_tl(filepath_tl)
        tl_warning_models = tl_data.get("warningModels", [])
        tl_lookup = tl_data.get("optionLookup", {})
        tl_visit_cycle = tl_data.get("visitCycleLookup", {})
        tl_model_display = tl_data.get("modelDisplayMap", {})
        tl_known_models = set(k.split('|')[0] for k in tl_lookup.keys())
    except Exception as e:
        print(f"[티엘] 파싱 실패: {e}")

    # ── 에이컴즈 파싱 ──
    print(f"[에이컴즈] 파싱 중: {filepath_sk}")
    data = parse_excel(filepath_sk)

    # ── E열 모델코드 정규화 (C열 모델코드 없는 제품) ──
    # WPUTD*114 → C형(하프), F형(스탠드) 두 제품으로 분리
    normalized_products = []
    for product in data["products"]:
        if product.get("modelCode"):
            normalized_products.append(product)
            continue

        # C열 모델코드 없는 제품: E열 option label에서 모델코드 추출
        # 첫 번째 옵션의 label이 E열 원본에서 파생됨
        # parse_excel에서 이미 clean_option_name 처리됨 → 원본 E열 필요
        # → 제품명을 기반으로 별도 Excel 재스캔
        # (간소화: products에 _raw_e 저장 방식 대신, 제품명→모델코드 매핑 사전 구축)
        normalized_products.append(product)

    # E열 원본값 재스캔으로 모델코드 추출
    import openpyxl as _openpyxl
    _wb = _openpyxl.load_workbook(filepath_sk, data_only=True)
    _ws = _wb.worksheets[0]
    _rows = list(_ws.iter_rows(values_only=True))
    DATA_START_IDX = 7

    # C열 없는 제품: 제품명 → {'codes': [...], 'raw_codes': [...], 'd_empty': bool} 매핑
    # d_empty=True: 제품 헤더 행에 D열도 비어있음 → 관리방식 상속 오염 → "방문관리" 강제
    _no_model_e_map = {}
    _cur_no_name = None
    for _row in _rows[DATA_START_IDX:]:
        _c2 = _row[2]
        _c3 = _row[3]   # D열
        _c4 = _row[4]
        if _c2 is not None and str(_c2).strip():
            _mc, _pname = parse_product_name_from_col2(_c2)
            if not _mc and _pname:
                _cur_no_name = _pname
                _e_val = clean(_c4) if _c4 else ""
                _d_val = clean(_c3) if _c3 else ""
                _raw_codes = extract_e_model_codes(_e_val)
                _matched = [tl_match_model(c, tl_known_models) for c in _raw_codes]
                _no_model_e_map[_cur_no_name] = {
                    "codes": _matched,
                    "raw_codes": _raw_codes,  # AK 원본 코드 (정규화 후)
                    "d_empty": not _d_val,
                }
            else:
                _cur_no_name = None

    # 모델코드 없는 제품에 모델코드 적용 + * 분리
    final_products = []
    code_mismatches = []  # 에이컴즈 ↔ 티엘 코드 상이 목록

    for product in data["products"]:
        if product.get("modelCode"):
            final_products.append(product)
            continue

        pname = product.get("name", "")
        info = _no_model_e_map.get(pname)

        if not info or not info["codes"]:
            final_products.append(product)
            continue

        codes = info["codes"]
        raw_codes = info.get("raw_codes", codes)
        for i, norm_code in enumerate(codes):
            display_code = tl_model_display.get(norm_code, norm_code)
            p_copy = dict(product)
            p_copy["options"] = [dict(o) for o in product["options"]]
            p_copy["modelCode"] = display_code
            p_copy["id"] = norm_code

            # 제품명에서 (하프/스탠드) 제거 후 형태 suffix 추가
            if len(codes) == 2:
                base_name = re.sub(r'\s*\(하프/스탠드\)\s*', '', pname).strip()
                p_copy["name"] = base_name + (" 하프형" if i == 0 else " 스탠드형")

            # D열이 비어있던 비MAT 제품만 관리방식 "방문관리" 고정
            # MAT 제품은 H열로 관리방식 이미 결정됨 → 덮어쓰지 않음
            if info["d_empty"] and p_copy.get("category") != "매트리스":
                for opt in p_copy["options"]:
                    opt["managementType"] = "방문관리"

            final_products.append(p_copy)

            # 코드 상이 감지: AK 원본 코드 ≠ TL 매칭 코드
            # MAT 제품은 시스템적 패턴(사이즈코드 차이)이므로 팝업 제외
            # RNW(AK 접미사) / RE(TL 접미사) 는 알려진 버전 패턴이므로 제외
            if i < len(raw_codes) and not norm_code.startswith('MAT'):
                raw = raw_codes[i]
                if raw != norm_code and raw != _norm_model(display_code):
                    if display_code and _fmt_code_display(raw) != display_code:
                        ak_n = _norm_model(_fmt_code_display(raw))
                        tl_n = _norm_model(display_code)
                        # 알려진 접미사 패턴: AK에 RNW, TL에 RE
                        is_known_suffix = (
                            ak_n.endswith('RNW') and ak_n[:-3] == tl_n
                        ) or (
                            tl_n.endswith('RE') and tl_n[:-2] == ak_n
                        )
                        if not is_known_suffix:
                            code_mismatches.append({
                                "name": p_copy["name"][:20],
                                "akCode": _fmt_code_display(raw),
                                "tlCode": display_code,
                            })

        codes_disp = [tl_model_display.get(c, c) for c in codes]
        msg = f"  [E열 정규화] {pname} -> {codes_disp}"
        print(msg.encode('cp949', errors='replace').decode('cp949'))

    data["products"] = final_products
    data["codeMismatches"] = code_mismatches

    # ── 접수처 비교 + 패키지 옵션 생성 ──
    def _mgmt_base_key(mgmt):
        if "방문" in mgmt: return "방문"
        if "셀프" in mgmt: return "셀프"
        if "관리없음" in mgmt: return "관리없음"
        return mgmt

    normalization_issues = []  # [{type, modelCode, name, akDetail, tlDetail, reason}]

    for product in data["products"]:
        model_code = product.get("modelCode", "")
        regular_opts = list(product["options"])
        package_opts = []
        seen_pkg = set()

        # ── 0. TL 제품 미리 찾기 (이후 전 단계에서 공통 사용) ──
        tl_products_data = tl_data.get("products", [])
        base_mvs = _tl_model_variants(model_code)
        all_mvs = _extend_model_variants_with_prefix(base_mvs, tl_known_models) if tl_known_models else base_mvs
        tl_prods_found = []
        for mv in all_mvs:
            tl_prods = [p for p in tl_products_data
                        if _norm_model(p["modelCode"]) == mv or
                        _norm_model(p["modelCode"]).startswith(mv) or
                        mv.startswith(_norm_model(p["modelCode"]))]
            if tl_prods:
                tl_prods_found = tl_prods
                break

        # ── 1. AK 빈 managementType → TL 교차참조로 보완 ──
        # 에이컴즈 D열에 관리방법이 없는 경우 TL 파일의 동일 약정 옵션에서 가져옴
        for opt in regular_opts:
            if opt.get("managementType") == "" and not opt.get("isPackage") and not opt.get("source"):
                for tl_prod in tl_prods_found:
                    for tl_opt in tl_prod.get("options", []):
                        if tl_opt["contractYears"] * 12 == opt["contractMonths"]:
                            opt["managementType"] = tl_opt["managementType"]
                            break
                    if opt.get("managementType"):
                        break

        # ── 1b. TL 교차참조 후에도 빈 경우 → 동일 제품 내 sibling 옵션에서 추론 ──
        # (예: 6년 AK 전용 약정 — TL에 없어서 못 채웠을 때)
        filled_mgmt_types = [
            o.get("managementType") for o in regular_opts
            if o.get("managementType") and not o.get("isPackage") and not o.get("source")
        ]
        if filled_mgmt_types:
            # 동일 제품의 채워진 관리방법 중 가장 많은 것 사용
            from collections import Counter
            dominant_mgmt = Counter(filled_mgmt_types).most_common(1)[0][0]
            for opt in regular_opts:
                if opt.get("managementType") == "" and not opt.get("isPackage") and not opt.get("source"):
                    opt["managementType"] = dominant_mgmt
                    opt["_mgmt_inferred"] = True

        # ── 2. 빈 visitCycle → 동일 제품 AK 옵션에서 보완 ──
        for opt in regular_opts:
            if not opt.get("visitCycle") and "방문" in (opt.get("managementType") or ""):
                visit = next(
                    (o.get("visitCycle", "") for o in regular_opts
                     if o.get("visitCycle") and "방문" in (o.get("managementType") or "")),
                    ""
                )
                if visit:
                    opt["visitCycle"] = visit

        # ── 3. 여전히 빈 managementType → 정규화 경고 ──
        missing_mgmt_opts = [o for o in regular_opts if not o.get("managementType") and not o.get("isPackage")]
        if missing_mgmt_opts:
            for o in missing_mgmt_opts:
                o["missingMgmt"] = True
            normalization_issues.append({
                "type": "MISSING_MGMT",
                "modelCode": model_code,
                "name": product.get("name", ""),
                "akDetail": "; ".join(
                    f'{o.get("contractLabel", "?")} 관리방법없음'
                    for o in missing_mgmt_opts[:3]
                ),
                "tlDetail": "에이컴즈·티엘 모두 관리방법 확인 불가",
                "reason": "관리방법을 특정할 수 없는 옵션 — 수동 확인 필요"
            })

        for opt in regular_opts:
            mgmt = opt.get("managementType", "")
            months = opt.get("contractMonths", 0)
            total_c = opt.get("totalCommission", 0)

            # 접수처 비교 (tl_known_models 전달 → prefix 매칭 활성화)
            _office_result = compute_recommended_office(
                tl_lookup, model_code, mgmt, months, total_c,
                is_package=False, tl_known_models=tl_known_models
            )
            opt["recommendedOffice"] = _office_result['office']
            # TL 수수료가 더 높으면 totalCommission을 TL 값으로 업데이트
            if _office_result['tlCommission'] is not None and _office_result['tlCommission'] > total_c:
                opt["totalCommission"] = _office_result['tlCommission']

            # 패키지 옵션 생성 (타사보상 제외, 중복 방지)
            if "타사보상" in mgmt:
                continue
            pkg_key = f"{mgmt}_{months}"
            if pkg_key in seen_pkg:
                continue
            seen_pkg.add(pkg_key)

            base_c  = opt.get("baseCommission", 0)
            add_c   = opt.get("additionalCommission", 0)
            bonus_c = opt.get("bonusCommission", 0)
            pkg_commission = round(base_c * 0.75) + add_c + bonus_c
            if pkg_commission <= 0:
                continue

            pkg_opt = dict(opt)
            pkg_opt["managementType"] = mgmt + "_패키지"
            pkg_opt["monthlyFee"] = max(0, opt["monthlyFee"] - 2000)
            pkg_opt["totalCommission"] = pkg_commission
            pkg_opt["isPackage"] = True
            _pkg_office_result = compute_recommended_office(
                tl_lookup, model_code, mgmt, months, pkg_commission,
                is_package=True, tl_known_models=tl_known_models
            )
            # TL 패키지 or 일반 수수료 비교 후 결정, 둘 다 없으면 에이컴즈
            pkg_opt["recommendedOffice"] = _pkg_office_result['office'] or '에이컴즈'
            if _pkg_office_result['tlCommission'] is not None and _pkg_office_result['tlCommission'] > pkg_commission:
                pkg_opt["totalCommission"] = _pkg_office_result['tlCommission']
            package_opts.append(pkg_opt)

        product["options"] = regular_opts + package_opts

        # ── TL 전용 관리방식 보완 ──
        # AK에 없는 관리방식이 TL에 있으면 TL 데이터로 합성 옵션 추가
        ak_mgmt_years = set(
            (_mgmt_base_key(o["managementType"]), o["contractMonths"] // 12)
            for o in product["options"]
            if not o.get("isPackage") and not o.get("source")
        )
        # 이미 추가된 TL보완 옵션 중복 방지
        tl_added = set()

        for tl_prod in tl_prods_found:
            for tl_opt in tl_prod.get("options", []):
                tl_mgmt = tl_opt["managementType"]
                tl_mgmt_base = _mgmt_base_key(tl_mgmt)
                tl_years = tl_opt["contractYears"]
                tl_months = tl_years * 12
                dedup_key = (tl_mgmt, tl_years)
                # AK에 이미 있는 관리방식+약정이면 스킵
                if (tl_mgmt_base, tl_years) in ak_mgmt_years:
                    continue
                # 이미 추가된 TL보완 중복 스킵
                if dedup_key in tl_added:
                    continue
                tl_added.add(dedup_key)
                # AK에 없는 TL 옵션 → 합성
                # visitCycle: AK 옵션 중 동일 관리방식 기반 것에서 참조
                ak_visit = next(
                    (o.get("visitCycle", "") for o in product["options"]
                     if _mgmt_base_key(o.get("managementType", "")) == tl_mgmt_base
                     and o.get("visitCycle")),
                    ""
                )
                syn_opt = {
                    "label": f"{tl_years}년",
                    "managementType": tl_mgmt,
                    "contractMonths": tl_months,
                    "contractLabel": f"{tl_years}년",
                    "monthlyFee": tl_opt["monthlyFee"],
                    "dataWarning": False,
                    "visitCycle": ak_visit,
                    "ownershipMonths": 0,
                    "registrationFee": 0,
                    "baseCommission": 0,
                    "additionalCount": 0,
                    "additionalCommission": 0,
                    "bonusCommission": 0,
                    "totalCommission": tl_opt["commission"],
                    "recommendedOffice": "티엘",
                    "source": "TL",
                }
                product["options"].append(syn_opt)
                msg = f"  [TL보완] {model_code} +{tl_mgmt} {tl_years}년"
                print(msg.encode('cp949', errors='replace').decode('cp949'))

        # 관리주기 정보 (TL G열 기준)
        for mv in all_mvs:
            if mv in tl_visit_cycle:
                product["visitCycleInfo"] = tl_visit_cycle[mv]
                break

        # ── 한쪽에만 있음(oneSideOnly) 판정 ──
        # non-패키지, non-TL보완 옵션 중 TL 매칭된 게 하나도 없으면 AK만 있는 제품
        regular_non_src = [o for o in product["options"]
                           if not o.get("isPackage") and not o.get("source")]
        matched_any = any(o.get("recommendedOffice") is not None for o in regular_non_src)
        if regular_non_src and not matched_any:
            product["oneSideOnly"] = "AK"
            # 정규화 이슈 기록
            ak_summary = "; ".join(
                f"{o.get('managementType','?')} {o.get('contractMonths',0)//12}년"
                for o in regular_non_src[:3]
            )
            normalization_issues.append({
                "type": "AK_ONLY",
                "modelCode": model_code,
                "name": product.get("name", ""),
                "akDetail": ak_summary,
                "tlDetail": "티엘 파일에 해당 제품 없음",
                "reason": "AK 수수료 파일에만 존재 — 티엘 파일에서 동일 모델 미확인"
            })

    # ── 정규화 이슈 추가: 남은 관리방식 미매칭 옵션(제품 일부만 매칭 실패) ──
    for product in data["products"]:
        if product.get("oneSideOnly"):
            continue  # 이미 위에서 처리됨
        regular_non_src = [o for o in product["options"]
                           if not o.get("isPackage") and not o.get("source")]
        unmatched = [o for o in regular_non_src if o.get("recommendedOffice") is None]
        matched = [o for o in regular_non_src if o.get("recommendedOffice") is not None]
        # 부분미매칭 옵션에 akOnlyOption 플래그 추가 → HTML 뱃지 표시용
        for o in unmatched:
            o["akOnlyOption"] = True
            o["recommendedOffice"] = "에이컴즈"  # AK만 있으면 접수처는 에이컴즈
        if unmatched and matched:
            # 일부 옵션만 매칭됨 → 부분 이슈로 기록
            ak_unmatched_summary = "; ".join(
                f"{o.get('managementType','?')} {o.get('contractMonths',0)//12}년"
                for o in unmatched[:3]
            )
            normalization_issues.append({
                "type": "PARTIAL",
                "modelCode": product.get("modelCode",""),
                "name": product.get("name",""),
                "akDetail": f"매칭 실패: {ak_unmatched_summary}",
                "tlDetail": "티엘 파일에 해당 약정/관리방식 없음",
                "reason": "일부 약정기간 또는 관리방식이 에이컴즈에만 존재"
            })

    # ── TL 전용 제품 감지 ──
    # AK 제품이 참조한 TL 모델코드 수집
    ak_referenced_tl = set()
    for product in data["products"]:
        mc = product.get("modelCode", "")
        if not mc:
            continue
        variants = _tl_model_variants(mc)
        extended = _extend_model_variants_with_prefix(variants, tl_known_models)
        for v in extended:
            ak_referenced_tl.add(v)

    for tl_product in tl_data.get("products", []):
        tl_norm = _norm_model(tl_product["modelCode"])
        if tl_norm in ak_referenced_tl:
            continue
        # TL 전용 제품 — products에 추가
        tl_only_entry = {
            "modelCode": tl_product["modelCode"],
            "name": tl_product["name"],
            "category": detect_category(tl_product["modelCode"], tl_product["name"], -1),
            "oneSideOnly": "TL",
            "options": [],
        }
        for tl_opt in tl_product["options"]:
            tl_years = tl_opt["contractYears"]
            tl_only_entry["options"].append({
                "label": f"{tl_years}년",
                "managementType": tl_opt["managementType"],
                "contractMonths": tl_years * 12,
                "contractYears": tl_years,
                "hasTasa": tl_opt.get("hasTasa", False),
                "monthlyFee": tl_opt["monthlyFee"],
                "baseCommission": 0,
                "bonusCommission1": 0,
                "bonusCommission2": 0,
                "totalCommission": tl_opt["commission"],
                "recommendedOffice": "티엘",
                "source": "TL",
                "dataWarning": False,
            })
        data["products"].append(tl_only_entry)
        normalization_issues.append({
            "type": "TL_ONLY",
            "modelCode": tl_product["modelCode"],
            "name": tl_product["name"],
            "akDetail": "에이컴즈 파일에 해당 제품 없음",
            "tlDetail": "; ".join(
                f"{o['managementType']} {o['contractYears']}년"
                for o in tl_product["options"][:3]
            ),
            "reason": "티엘 수수료 파일에만 존재 — 에이컴즈에서 해당 제품 미확인",
        })
        msg = f"  [TL전용] {tl_product['modelCode']} — {tl_product['name']}"
        print(msg.encode('cp949', errors='replace').decode('cp949'))

    # ── TL 사업자전용 옵션 주입 ──
    tl_biz = tl_data.get("bizOptions", [])
    if tl_biz:
        for product in data["products"]:
            mc_norm = _norm_model(product.get("modelCode", ""))
            base_mvs = _tl_model_variants(product.get("modelCode", ""))
            all_mvs_biz = _extend_model_variants_with_prefix(base_mvs, tl_known_models) if tl_known_models else base_mvs
            for biz in tl_biz:
                if biz["normCode"] not in all_mvs_biz and _norm_model(biz["modelCode"]) not in all_mvs_biz:
                    continue
                visit = biz.get("visitCycle") or ""
                if not visit:
                    visit = product.get("visitCycleInfo", "")
                product["options"].append({
                    "label":            f"{biz['contractYears']}년",
                    "managementType":   biz["managementType"],
                    "contractMonths":   biz["contractYears"] * 12,
                    "contractLabel":    biz["contractLabel"],
                    "monthlyFee":       biz["monthlyFee"],
                    "totalCommission":  biz["commission"],
                    "recommendedOffice": "티엘",
                    "source":           "TL",
                    "isBizOnly":        True,
                    "bizDiscount":      f"{biz['discountPct']}%",
                    "obligation":       biz.get("obligation", ""),
                    "visitCycle":       visit,
                    "isPackage":        False,
                    "dataWarning":      False,
                    "baseCommission":   0,
                    "bonusCommission":  0,
                    "registrationFee":  0,
                })

    data["normalizationIssues"] = normalization_issues

    # ── 옵션 약정 오름차순 정렬 (신규→타사보상→패키지, 각 그룹 내 contractMonths 오름차순) ──
    def _sort_key(o):
        is_biz = o.get("isBizOnly", False)
        is_pkg = o.get("isPackage", False)
        mgmt = o.get("managementType") or ""
        is_tasa = "타사보상" in mgmt
        # 그룹: 0=일반, 1=타사보상, 2=패키지, 3=사업자전용
        group = 3 if is_biz else (2 if is_pkg else (1 if is_tasa else 0))
        disc = o.get("bizDiscount", "0%").replace("%", "")
        try:
            disc_int = int(disc)
        except ValueError:
            disc_int = 0
        return (group, o.get("contractMonths", 0), disc_int)

    for product in data["products"]:
        product["options"].sort(key=_sort_key)

    # ── JSON 저장 ──
    json_out = os.path.join(base_dir, "sk_data.json")
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"JSON 저장: {json_out}")

    print(f"정규화 이슈: {len(normalization_issues)}건 (AK만: {sum(1 for x in normalization_issues if x['type']=='AK_ONLY')}건, TL만: {sum(1 for x in normalization_issues if x['type']=='TL_ONLY')}건, 부분: {sum(1 for x in normalization_issues if x['type']=='PARTIAL')}건)")

    # ── HTML 생성 ──
    tpl_path = os.path.join(base_dir, "렌탈수수료_템플릿.html")
    if os.path.exists(tpl_path):
        with open(tpl_path, "r", encoding="utf-8") as f:
            html = f.read()

        sk_js   = json.dumps(data, ensure_ascii=False)
        tl_js   = json.dumps(tl_warning_models, ensure_ascii=False)
        cm_js   = json.dumps(code_mismatches, ensure_ascii=False)
        ni_js   = json.dumps(normalization_issues, ensure_ascii=False)

        # LG 데이터 로드 (lg_data.json, lg_air_data.json)
        lg_json_path     = os.path.join(base_dir, "lg_data.json")
        lg_air_json_path = os.path.join(base_dir, "lg_air_data.json")
        lg_js     = "{}"
        lg_air_js = "{}"
        lg_water_norm_js = "[]"
        lg_air_norm_js   = "[]"
        if os.path.exists(lg_json_path):
            with open(lg_json_path, "r", encoding="utf-8") as f:
                lg_raw = json.load(f)
            lg_js = json.dumps(lg_raw, ensure_ascii=False)
            lg_water_norm_js = json.dumps(lg_raw.get("normalizationIssues", []), ensure_ascii=False)
            print(f"LG 정수기 데이터 로드: {len(lg_raw.get('products', []))}개 제품")
        else:
            print(f"[경고] LG 정수기 JSON 없음: {lg_json_path}")
        if os.path.exists(lg_air_json_path):
            with open(lg_air_json_path, "r", encoding="utf-8") as f:
                lg_air_raw = json.load(f)
            lg_air_js = json.dumps(lg_air_raw, ensure_ascii=False)
            lg_air_norm_js = json.dumps(lg_air_raw.get("normalizationIssues", []), ensure_ascii=False)
            print(f"LG 공청기 데이터 로드: {len(lg_air_raw.get('products', []))}개 제품")
        else:
            print(f"[경고] LG 공청기 JSON 없음: {lg_air_json_path}")

        # ── 파싱 리포트 생성 ──
        cats = {}
        for p in data["products"]:
            cats[p["category"]] = cats.get(p["category"], 0) + 1
        biz_count = sum(1 for p in data["products"] for o in p["options"] if o.get("isBizOnly"))
        dw_count  = sum(1 for p in data["products"] for o in p["options"] if o.get("dataWarning"))
        tl_products = tl_data.get("products", [])
        tl_merged   = tl_data.get("mergedVariants", [])
        tl_biz      = tl_data.get("bizOptions", [])
        ni_total = len(normalization_issues)
        parse_report = {
            "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "sk": {
                "sourceFile": os.path.basename(filepath_sk),
                "totalProducts": len(data["products"]),
                "byCategory": dict(sorted(cats.items())),
                "bizOptionsCount": biz_count,
                "dataWarningCount": dw_count,
            },
            "tl": {
                "sourceFile": os.path.basename(filepath_tl),
                "parsedProducts": len(tl_products),
                "mergedVariants": len(tl_merged),
                "bizOptionsCount": len(tl_biz),
            },
            "normalizationIssues": {
                "total": ni_total,
                "akOnly": sum(1 for x in normalization_issues if x["type"] == "AK_ONLY"),
                "tlOnly": sum(1 for x in normalization_issues if x["type"] == "TL_ONLY"),
                "partial": sum(1 for x in normalization_issues if x["type"] == "PARTIAL"),
            },
            "codeMismatches": len(code_mismatches),
            "lg": {
                "waterSourceFile": os.path.basename(lg_json_path) if os.path.exists(lg_json_path) else None,
                "waterProducts": len(lg_raw.get("products", [])) if os.path.exists(lg_json_path) else 0,
                "airSourceFile": os.path.basename(lg_air_json_path) if os.path.exists(lg_air_json_path) else None,
                "airProducts": len(lg_air_raw.get("products", [])) if os.path.exists(lg_air_json_path) else 0,
            },
        }
        pr_js = json.dumps(parse_report, ensure_ascii=False)

        html_out_str = html.replace("__SK_DATA__", sk_js) \
                           .replace("__TL_WARNINGS__", tl_js) \
                           .replace("__CODE_MISMATCHES__", cm_js) \
                           .replace("__NORM_ISSUES__", ni_js) \
                           .replace("__LG_DATA__", lg_js) \
                           .replace("__LG_AIR_DATA__", lg_air_js) \
                           .replace("__LG_WATER_NORM_ISSUES__", lg_water_norm_js) \
                           .replace("__LG_AIR_NORM_ISSUES__", lg_air_norm_js) \
                           .replace("__PARSE_REPORT__", pr_js)

        month_tag = data["metadata"].get("parsedAt", "")[:7].replace("-","")[2:]
        out_html = os.path.join(base_dir, f"렌탈수수료_{month_tag}.html")
        with open(out_html, "w", encoding="utf-8") as f:
            f.write(html_out_str)
        print(f"HTML 저장: {out_html}")

    print(f"\n완료! 제품 수: {len(data['products'])}개")
    cats = {}
    for p in data['products']:
        cats[p['category']] = cats.get(p['category'], 0) + 1
    for cat, cnt in sorted(cats.items()):
        print(f"  {cat}: {cnt}개")
