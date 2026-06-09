import os
import json
import pickle
from pathlib import Path

from google.oauth2.credentials import Credentials
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
import gspread
import openpyxl

# ─────────────────────────── 설정 ───────────────────────────
SPREADSHEET_ID  = '1y5wfMhcM3_S7FnJWHAhTIkqyzFZZCcItGQpVL5nu5XY'
EXCEL_PATH      = Path(r'C:\Users\a\Desktop\안티그라비티\입금명단자동화\신청현황_자동화테스트버전.xlsx')
SCOPES          = ['https://www.googleapis.com/auth/spreadsheets.readonly']
BASE_DIR        = Path(__file__).parent
TOKEN_PATH      = BASE_DIR / 'token.pickle'
CREDENTIALS_PATH = BASE_DIR / 'credentials.json'
BACKUP_MAX      = 30
DATA_START_ROW  = 3   # 데이터 시작 행 (2행은 고정행)

TELECOM_MAP = {
    'KT': 'KT', 'SKT': 'SK', 'SKB': 'SK', 'SK알뜰': 'SK',
    'LGU+': 'LG', 'LG소호': 'LG', '스카이': '스카이', 'LG헬로': 'LG헬로',
}
PRODUCT_MAP = {'인단': '인터넷', '번들': '인터넷+TV'}


# ─────────────────────────── 이름 마스킹 ───────────────────────────
def mask_name(name: str) -> str:
    name = name.strip()
    n = len(name)
    if n <= 1:
        return name
    if n == 2:
        return name[0] + '*'
    return name[0] + '*' * (n - 2) + name[-1]


# ─────────────────────────── Google 인증 ───────────────────────────
def get_credentials():
    creds = None

    if TOKEN_PATH.exists():
        with open(TOKEN_PATH, 'rb') as f:
            creds = pickle.load(f)

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not CREDENTIALS_PATH.exists():
                raise FileNotFoundError(
                    f"credentials.json 파일이 없습니다.\n"
                    f"Google Cloud Console에서 OAuth 클라이언트 ID를 생성한 후\n"
                    f"{CREDENTIALS_PATH} 위치에 저장해주세요."
                )
            flow = InstalledAppFlow.from_client_secrets_file(str(CREDENTIALS_PATH), SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_PATH, 'wb') as f:
            pickle.dump(creds, f)

    return creds


# ─────────────────────────── 구글시트 데이터 추출 ───────────────────────────
def fetch_sheets():
    client = gspread.authorize(get_credentials())
    ss = client.open_by_key(SPREADSHEET_ID)
    wired  = ss.worksheet('유선개통').get_all_values()
    rental = ss.worksheet('렌탈개통').get_all_values()
    return wired, rental


def process_wired(data):
    """유선개통 탭 → (고객명, 통신사, 상품, 은행) 목록"""
    rows = []
    for row in data[1:]:   # 1행 헤더 제외
        if len(row) < 12:
            continue
        telecom = row[2].strip()    # C열
        name    = row[6].strip()    # G열
        bank    = row[7].strip()    # H열
        product = row[11].strip()   # L열

        if not name:
            continue
        if '유심' in product:
            continue

        rows.append((
            mask_name(name),
            TELECOM_MAP.get(telecom, telecom),
            PRODUCT_MAP.get(product, product),
            bank,
        ))
    return rows


def process_rental(data):
    """렌탈개통 탭 → (고객명, '', '가전렌탈', 은행) 목록"""
    rows = []
    for row in data[1:]:
        if len(row) < 8:
            continue
        name = row[6].strip()   # G열
        bank = row[7].strip()   # H열

        if not name:
            continue

        rows.append((mask_name(name), '', '가전렌탈', bank))
    return rows


# ─────────────────────────── 엑셀 업데이트 ───────────────────────────
def update_excel(wired_rows, rental_rows):
    if EXCEL_PATH.exists():
        wb = openpyxl.load_workbook(EXCEL_PATH)
    else:
        print(f"  엑셀 파일이 없어 새로 생성합니다: {EXCEL_PATH}")
        wb = openpyxl.Workbook()

    ws = wb.active

    # A열 비어있고 D열에 값 있는 행 = 백업 데이터
    backup = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        a_val = ws.cell(r, 1).value
        d_val = ws.cell(r, 4).value
        a_empty = (a_val is None or str(a_val).strip() == '')
        if a_empty and d_val:
            backup.append((
                ws.cell(r, 4).value,  # D 고객명
                ws.cell(r, 5).value,  # E 통신사
                ws.cell(r, 6).value,  # F 상품명
                ws.cell(r, 7).value,  # G 은행
            ))

    backup = backup[:BACKUP_MAX]

    # D~G열 초기화 (3행~)
    for r in range(DATA_START_ROW, ws.max_row + 1):
        for c in (4, 5, 6, 7):
            ws.cell(r, c).value = None

    # 전체 = 백업 + 유선개통 + 렌탈개통, D열 기준 가나다 오름차순
    all_rows = backup + wired_rows + rental_rows
    all_rows.sort(key=lambda x: str(x[0] or ''))

    for i, (d, e, f, g) in enumerate(all_rows):
        r = DATA_START_ROW + i
        ws.cell(r, 4).value = d
        ws.cell(r, 5).value = e or None
        ws.cell(r, 6).value = f
        ws.cell(r, 7).value = g

    wb.save(EXCEL_PATH)

    print(f"  저장 완료 → {EXCEL_PATH}")
    print(f"  백업: {len(backup)}행 | 유선개통: {len(wired_rows)}행 | 렌탈개통: {len(rental_rows)}행 | 합계: {len(all_rows)}행")


# ─────────────────────────── 실행 ───────────────────────────
def main():
    print("=" * 50)
    print("입금명단 자동화 — 1단계: 구글시트 → 엑셀")
    print("=" * 50)

    print("\n[1] 구글시트 데이터 가져오는 중...")
    wired_data, rental_data = fetch_sheets()

    print("[2] 데이터 가공 중...")
    wired_rows  = process_wired(wired_data)
    rental_rows = process_rental(rental_data)
    print(f"  유선개통: {len(wired_rows)}건 / 렌탈개통: {len(rental_rows)}건")

    print("[3] 엑셀 업데이트 중...")
    update_excel(wired_rows, rental_rows)

    print("\n완료!")


if __name__ == '__main__':
    main()
